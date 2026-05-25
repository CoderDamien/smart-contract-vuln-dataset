from __future__ import annotations

import csv
import hashlib
import json
import re
import shutil
from pathlib import Path

from openpyxl import load_workbook


SOURCE_ROOT = Path(__file__).resolve().parents[3]
RELEASE_ROOT = Path(__file__).resolve().parents[1]

PAPER_DIR = SOURCE_ROOT / "papers" / "当前论文"
PAPER_XLSX = PAPER_DIR / "本地实验结果汇总_按模型任务结果.xlsx"
SPLIT_ROOT = SOURCE_ROOT / "data" / "prepared" / "balanced_stage1_resplit_721"
REMOTE_RESULTS = SOURCE_ROOT / "results" / "remote_mirror"

REPLICATION_ROOT = RELEASE_ROOT / "replication"


SHEETS = {
    "paper_experiment_matrix_closure.csv": "论文实验矩阵闭合明细",
    "paper_experiment_metrics_all.csv": "论文实验指标总表",
    "paper_experiment_closure_summary.csv": "论文实验闭合汇总",
    "all_result_metrics.csv": "全部结果指标",
    "has_vul_metrics.csv": "has指标",
    "vul_type_metrics.csv": "type指标",
    "vul_line_metrics.csv": "line指标",
    "prompt_ablation.csv": "Prompt消融",
    "planned_matrix.csv": "计划矩阵",
    "run_closure_audit.csv": "run闭合审计",
    "data_completion_before_after.csv": "补齐前后12条明细",
    "data_completion_pairs.csv": "补齐前后配对",
    "model_summary.csv": "按模型汇总",
}

TASK_DIRS = {
    "has_vul": "has_vul_721_stratified_v1",
    "vul_type": "vul_type_721_stratified_v1",
    "vul_line": "vul_line_721_stratified_v1",
}

LABELS = [
    "access_control",
    "arithmetic",
    "bad_randomness",
    "denial_service",
    "front_running",
    "reentrancy",
    "time_manipulation",
    "unchecked_low_calls",
]


def reset_dir(path: Path) -> None:
    if path.exists():
        shutil.rmtree(path)
    path.mkdir(parents=True, exist_ok=True)


def copy_file(src: Path, dst: Path) -> None:
    dst.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(src, dst)


def sha256(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as fh:
        for chunk in iter(lambda: fh.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def safe_name(text: str) -> str:
    text = re.sub(r"[^A-Za-z0-9._-]+", "_", text.strip())
    text = re.sub(r"_+", "_", text).strip("._")
    return text[:180] or "unnamed"


def write_json(path: Path, obj) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(obj, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def export_sheet(ws, out_path: Path) -> list[dict[str, str]]:
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(x).strip() if x is not None else "" for x in rows[0]]
    data: list[dict[str, str]] = []
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with out_path.open("w", newline="", encoding="utf-8-sig") as fh:
        writer = csv.DictWriter(fh, fieldnames=headers)
        writer.writeheader()
        for row in rows[1:]:
            item = {
                header: "" if value is None else str(value)
                for header, value in zip(headers, row)
            }
            data.append(item)
            writer.writerow(item)
    return data


def export_excel_tables() -> dict[str, list[dict[str, str]]]:
    wb = load_workbook(PAPER_XLSX, read_only=True, data_only=True)
    exported: dict[str, list[dict[str, str]]] = {}
    for filename, sheet_name in SHEETS.items():
        exported[filename] = export_sheet(wb[sheet_name], REPLICATION_ROOT / "results" / filename)
    copy_file(PAPER_XLSX, REPLICATION_ROOT / "results" / "paper_experiment_results.xlsx")
    return exported


def export_splits() -> None:
    split_out = REPLICATION_ROOT / "data_splits" / "balanced_stage1_resplit_721"
    records = []
    for task, dirname in TASK_DIRS.items():
        for split in ("train", "val", "test"):
            src = SPLIT_ROOT / dirname / f"{split}.json"
            records.append(
                {
                    "task": task,
                    "split": split,
                    "dataset_archive": "data/processed/balanced_stage1_resplit_721.tar.gz",
                    "archive_internal_path": f"balanced_stage1_resplit_721/{dirname}/{split}.json",
                    "source_size_bytes": src.stat().st_size,
                    "sha256": sha256(src),
                }
            )
    write_csv(split_out / "split_file_index.csv", records)
    copy_file(SPLIT_ROOT / "build_report.json", split_out / "build_report.json")


def export_label_mapping() -> None:
    config_src = SOURCE_ROOT / "config" / "dataset_label_mapping_draft.json"
    copy_file(config_src, REPLICATION_ROOT / "label_mapping" / "dataset_label_mapping_draft.json")
    write_json(
        REPLICATION_ROOT / "label_mapping" / "paper_label_space.json",
        {
            "version": "paper-release-v1",
            "task_labels": LABELS,
            "notes": [
                "The paper reports the eight normalized vulnerability categories listed in task_labels.",
                "Labels mapped to other, unknown, ignore, or safe are not part of the main vul_type label space.",
            ],
        },
    )


def export_scripts_and_configs() -> None:
    for src, dst in [
        (SOURCE_ROOT / "src" / "eval" / "metrics_vul_line.py", "evaluation/metrics_vul_line.py"),
        (SOURCE_ROOT / "src" / "tasks" / "has_vul.py", "evaluation/tasks/has_vul.py"),
        (SOURCE_ROOT / "src" / "tasks" / "vul_type.py", "evaluation/tasks/vul_type.py"),
        (SOURCE_ROOT / "src" / "tasks" / "vul_line.py", "evaluation/tasks/vul_line.py"),
        (SOURCE_ROOT / "src" / "prompts" / "templates.py", "evaluation/prompts/templates.py"),
        (SOURCE_ROOT / "src" / "prompts" / "parsers.py", "evaluation/prompts/parsers.py"),
        (SOURCE_ROOT / "scripts" / "generate_q1_paper_assets.py", "figure_source/generate_q1_paper_assets.py"),
        (SOURCE_ROOT / "scripts" / "redraw_q1_figures.py", "figure_source/redraw_q1_figures.py"),
        (SOURCE_ROOT / "docs" / "formal_experiment_execution_plan.md", "experiment_configs/formal_experiment_execution_plan.md"),
        (SOURCE_ROOT / "scripts" / "runtime" / "run_formal_balanced721_queue.sh", "experiment_configs/run_formal_balanced721_queue.sh"),
        (SOURCE_ROOT / "scripts" / "runtime" / "enrich_matrix_runtime.py", "experiment_configs/enrich_matrix_runtime.py"),
    ]:
        copy_file(src, REPLICATION_ROOT / dst)
    config_out = REPLICATION_ROOT / "experiment_configs" / "project_configs"
    for src in (SOURCE_ROOT / "config").glob("*.json"):
        copy_file(src, config_out / src.name)


def export_figures() -> None:
    for fig_dir in [PAPER_DIR / "figures_q1", PAPER_DIR / "figures_q1_redesign"]:
        if not fig_dir.exists():
            continue
        target = REPLICATION_ROOT / "figures" / fig_dir.name
        target.mkdir(parents=True, exist_ok=True)
        for src in fig_dir.glob("*.svg"):
            copy_file(src, target / src.name)
        notes = fig_dir / "figure_redesign_notes.md"
        if notes.exists():
            copy_file(notes, target / notes.name)


def index_prediction_artifacts(rows: list[dict[str, str]]) -> None:
    run_names = sorted({r.get("run_name", "").strip() for r in rows if r.get("run_name", "").strip()})
    by_run: dict[str, dict[str, list[Path]]] = {}
    for run_name in run_names:
        by_run[run_name] = {"config": [], "metrics": [], "state": [], "predictions": []}

    interesting = set(run_names)
    for path in REMOTE_RESULTS.rglob("*"):
        if not path.is_file():
            continue
        text = str(path)
        matched = next((run for run in interesting if run in text), None)
        if not matched:
            continue
        name = path.name
        if name == "config.json":
            by_run[matched]["config"].append(path)
        elif name == "metrics.json":
            by_run[matched]["metrics"].append(path)
        elif name == "state.json":
            by_run[matched]["state"].append(path)
        elif name.startswith("predictions") and name.endswith(".jsonl"):
            by_run[matched]["predictions"].append(path)

    config_index = []
    config_bundle = []
    prediction_index = []
    for row in rows:
        run_name = row.get("run_name", "").strip()
        if not run_name:
            continue
        task = row.get("任务") or row.get("task") or ""
        method = row.get("方法") or row.get("method") or ""
        model = row.get("模型") or row.get("model") or ""
        machine = row.get("执行机器") or ""
        bundle = by_run.get(run_name, {})

        for src in bundle.get("config", [])[:1]:
            raw = src.read_text(encoding="utf-8", errors="replace")
            config_index.append(
                {
                    "run_name": run_name,
                    "model": model,
                    "task": task,
                    "method": method,
                    "machine": machine,
                    "config_bundle": "experiment_configs/run_configs_bundle.jsonl",
                    "config_record_id": safe_name(run_name),
                    "config_sha256": hashlib.sha256(raw.encode("utf-8")).hexdigest(),
                    "source_path": str(src.relative_to(SOURCE_ROOT)),
                }
            )
            config_bundle.append(
                {
                    "record_id": safe_name(run_name),
                    "run_name": run_name,
                    "model": model,
                    "task": task,
                    "method": method,
                    "machine": machine,
                    "source_path": str(src.relative_to(SOURCE_ROOT)),
                    "config": json.loads(raw),
                }
            )

        for src in bundle.get("predictions", []):
            try:
                line_count = sum(1 for _ in src.open("r", encoding="utf-8", errors="ignore"))
            except OSError:
                line_count = ""
            prediction_index.append(
                {
                    "run_name": run_name,
                    "model": model,
                    "task": task,
                    "method": method,
                    "machine": machine,
                    "prediction_file_name": src.name,
                    "prediction_rows": line_count,
                    "size_bytes": src.stat().st_size,
                    "sha256": sha256(src),
                    "source_path": str(src.relative_to(SOURCE_ROOT)),
                }
            )

    write_csv(REPLICATION_ROOT / "experiment_configs" / "run_configs_index.csv", config_index)
    bundle_path = REPLICATION_ROOT / "experiment_configs" / "run_configs_bundle.jsonl"
    bundle_path.parent.mkdir(parents=True, exist_ok=True)
    with bundle_path.open("w", encoding="utf-8") as fh:
        for item in config_bundle:
            fh.write(json.dumps(item, ensure_ascii=False, sort_keys=True) + "\n")
    write_csv(REPLICATION_ROOT / "prediction_summaries" / "prediction_artifact_index.csv", prediction_index)
    write_json(REPLICATION_ROOT / "prediction_summaries" / "prediction_artifact_index.json", prediction_index)


def write_csv(path: Path, rows: list[dict]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    fieldnames = sorted({key for row in rows for key in row.keys()})
    with path.open("w", newline="", encoding="utf-8-sig") as fh:
        writer = csv.DictWriter(fh, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def write_manifest() -> None:
    files = []
    for path in sorted(REPLICATION_ROOT.rglob("*")):
        if path.is_file():
            files.append(
                {
                    "path": path.relative_to(RELEASE_ROOT).as_posix(),
                    "size_bytes": path.stat().st_size,
                    "sha256": sha256(path),
                }
            )
    write_json(REPLICATION_ROOT / "manifest.json", {"file_count": len(files), "files": files})


def main() -> None:
    reset_dir(REPLICATION_ROOT)
    export_splits()
    export_label_mapping()
    export_scripts_and_configs()
    exported = export_excel_tables()
    export_figures()
    index_prediction_artifacts(exported["paper_experiment_metrics_all.csv"])
    copy_file(SOURCE_ROOT / "results" / "remote_mirror" / "local_completed_results_tables.md", REPLICATION_ROOT / "results" / "local_completed_results_tables.md")
    write_manifest()


if __name__ == "__main__":
    main()
