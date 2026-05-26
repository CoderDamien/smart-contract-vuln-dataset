from __future__ import annotations

import csv
import hashlib
import json
import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
REPLICATION = ROOT / "replication"
RESULTS = REPLICATION / "results"
SUPP = ROOT / "supplementary"
TABLES = SUPP / "tables"


TABLE_DEFS = [
    ("S1", "Formal experiment matrix and completion status", "paper_experiment_matrix_closure.csv"),
    ("S2", "Model inventory and experiment coverage by model family", "model_summary.csv"),
    ("S3", "Complete model-task-method metric matrix", "paper_experiment_metrics_all.csv"),
    ("S4", "Vulnerability existence detection results", "has_vul_metrics.csv"),
    ("S5", "Vulnerability type classification results", "vul_type_metrics.csv"),
    ("S6", "Vulnerable line localization results", "vul_line_metrics.csv"),
    ("S7", "Traditional pretrained-code baselines", None),
    ("S8", "Qwen2.5-Coder parameter-scale analysis", None),
    ("S9", "Prompt ablation results for Qwen2.5-Coder-7B and 32B", "prompt_ablation.csv"),
    ("S10", "Dataset expansion before/after paired results", "data_completion_pairs.csv"),
    ("S11", "Complete method-pair comparison records", None),
    ("S12", "Runtime and performance trade-off records", "all_result_metrics.csv"),
    ("S13", "Prediction artifact and parsing-error audit", None),
    ("S14", "Figure source data and generated figure inventory", None),
]

ZH_TITLES = {
    "Table S1": "正式实验矩阵与完成状态",
    "Table S2": "模型清单与模型族实验覆盖",
    "Table S3": "完整模型-任务-方法指标矩阵",
    "Table S4": "漏洞存在性检测结果",
    "Table S5": "漏洞类型分类结果",
    "Table S6": "漏洞行定位结果",
    "Table S7": "传统预训练代码模型基线",
    "Table S8": "Qwen2.5-Coder 参数规模分析",
    "Table S9": "Qwen2.5-Coder-7B 与 32B 的提示消融结果",
    "Table S10": "数据扩展前后配对结果",
    "Table S11": "完整方法配对比较记录",
    "Table S12": "运行成本与性能权衡记录",
    "Table S13": "预测产物与解析错误审计",
    "Table S14": "图表源数据与生成图文件清单",
}

KEY_COLUMNS = {
    "Table S1": "model; task; method; status; execution_source",
    "Table S2": "model_family; model; task_count; completed_count",
    "Table S3": "model; task; method; status; f1/macro_f1/tolerant_f1",
    "Table S4": "model; method; f1; precision; recall",
    "Table S5": "model; method; macro_f1; multi_label_f1; status",
    "Table S6": "model; method; strict_f1; tolerant_f1; contract_hit",
    "Table S7": "model; task; method; status; f1/macro_f1/tolerant_f1",
    "Table S8": "model; task; method; status; model_family",
    "Table S9": "model; task; prompt_version; mode; parseable_ratio; f1/strict_f1/tolerant_f1",
    "Table S10": "model; task; method; expansion_stage; dataset_view",
    "Table S11": "model; task; comparison; left_metric; right_metric; delta",
    "Table S12": "model; task; method; runtime_minutes; status",
    "Table S13": "source; run_name; model; task; method; parseable_ratio",
    "Table S14": "artifact_type; path; sha256",
}

HEADER_MAP = {
    "系列": "model_family",
    "模型": "model",
    "任务": "task",
    "方法": "method",
    "任务类型": "task_type",
    "状态": "status",
    "结果状态": "result_status",
    "计划标记": "plan_marker",
    "执行机器": "execution_source",
    "机器/来源": "execution_source",
    "来源": "source",
    "本地metrics路径": "metrics_path",
    "本地summary": "summary_path",
    "完成时间": "completed_time",
    "备注": "notes",
    "说明": "notes",
    "本地metrics闭合状态": "metrics_closure_status",
    "闭合审计说明": "closure_audit_notes",
    "补齐阶段说明": "expansion_stage",
    "数据版本说明": "dataset_view",
    "补齐前6组标记": "pre_expansion_pair_marker",
    "可用于补齐前后对比": "paired_expansion_counterpart",
    "数据集": "dataset_view",
    "Prompt模板": "prompt_template",
    "版本": "prompt_version",
    "模式": "mode",
    "补齐前metrics状态": "pre_expansion_metrics_status",
    "补齐后metrics状态": "post_expansion_metrics_status",
    "补齐前run（原始3000条）": "pre_expansion_run",
    "补齐后run（常规大体量）": "post_expansion_run",
    "配对状态": "pairing_status",
    "未完成任务名称": "incomplete_task_names",
    "应跑总数": "planned_total",
    "不支持/需适配": "unsupported_or_requires_adaptation",
    "失败": "failed",
    "已完成": "completed",
    "缺失": "missing",
    "运行中": "running",
    "阻塞": "blocked",
}

VALUE_MAP = {
    "补齐前后对比": "dataset_expansion_comparison",
    "补齐前": "before_expansion",
    "补齐后": "after_expansion",
    "原始3000条": "SmartBugs Curated early view",
    "原始3000条数据集": "SmartBugs Curated early view",
    "常规实验大体量数据（约24000+24000+12000）": "multi-source formal view (about 24k/24k/12k)",
    "是：原始3000条补齐前6组之一": "yes: one of the six pre-expansion paired runs",
    "否：这是补齐后侧，不是原始3000条前侧": "no: post-expansion counterpart",
    "已闭合": "closed",
    "有结果": "completed",
    "已完成": "completed",
    "本地镜像缺失": "local_mirror_missing",
    "运行中": "running",
    "传统基准": "traditional_baseline",
    "Prompt消融": "prompt_ablation",
    "旧版3B prompt消融": "legacy_3b_prompt_ablation",
    "完整配对": "complete_pair",
    "常规实验": "formal_experiment",
    "缺失": "missing",
    "失败": "failed",
    "阻塞": "blocked",
    "不支持/需适配": "unsupported_or_requires_adaptation",
    "SmartBugs Curated early view相关（非本次6组）": "SmartBugs Curated early-view related (not one of the six paired runs)",
    "否：SmartBugs Curated early view相关结果，但不是本次before_expansion6组": "no: SmartBugs Curated early-view result, but not one of the six pre-expansion paired runs",
    "否；不要放入本次6组before_expansion后主对比": "no: exclude from the main six-pair pre/post expansion comparison",
    "补齐前run（原始3000条）": "pre_expansion_run",
    "补齐后run（常规大体量）": "post_expansion_run",
    "补齐前metrics状态": "pre_expansion_metrics_status",
    "补齐后metrics状态": "post_expansion_metrics_status",
    "配对状态": "pairing_status",
    "应跑总数": "planned_total",
    "未完成任务名称": "incomplete_task_names",
}


def read_csv(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as fh:
        return list(csv.DictReader(fh))


def write_csv(path: Path, rows: list[dict[str, str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    fields = sorted({key for row in rows for key in row.keys()})
    if not fields:
        fields = ["note"]
        rows = [{"note": "No records"}]
    with path.open("w", encoding="utf-8-sig", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=fields)
        writer.writeheader()
        writer.writerows(rows)


def normalize_cell(text: str) -> str:
    if text is None:
        return ""
    text = str(text)
    if text in VALUE_MAP:
        return VALUE_MAP[text]
    for zh, en in VALUE_MAP.items():
        if zh in text:
            text = text.replace(zh, en)
    text = text.replace("同方法的补齐后侧配对", "paired post-expansion counterpart")
    text = text.replace("同方法的补齐前侧配对", "paired pre-expansion counterpart")
    text = text.replace("同方法的after_expansion侧配对", "paired post-expansion counterpart")
    text = text.replace("同方法的before_expansion侧配对", "paired pre-expansion counterpart")
    text = text.replace("同基线的after_expansion结果配对", "paired post-expansion baseline result")
    text = text.replace("历史对照", "historical_reference")
    text = text.replace("after_expansion大体量数据run；已从“before_expansion后12条明细”同步修正", "post-expansion large-scale dataset run; corrected from pre/post expansion detail records")
    text = text.replace("B机未保留OSS checkpoint，D机重新补齐完成", "compute_node_b did not retain the OSS checkpoint; compute_node_d reran and completed the missing run")
    text = text.replace("predictions齐全", "predictions_complete")
    text = text.replace("；", "; ")
    text = text.replace("。", ".")
    text = text.replace("“", "\"").replace("”", "\"")
    return text


def normalize_rows(rows: list[dict[str, str]]) -> list[dict[str, str]]:
    normalized = []
    for row in rows:
        out: dict[str, str] = {}
        for key, val in row.items():
            new_key = HEADER_MAP.get(key, key)
            new_val = normalize_cell(val)
            if new_key in out and out[new_key] and new_val and out[new_key] != new_val:
                out[new_key] = f"{out[new_key]}; {new_val}"
            elif new_val or new_key not in out:
                out[new_key] = new_val
        normalized.append(out)
    return normalized


def sha256(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as fh:
        for chunk in iter(lambda: fh.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def value(row: dict[str, str], *names: str) -> str:
    for name in names:
        if name in row and str(row[name]).strip():
            return str(row[name]).strip()
    return ""


def number(text: str):
    if text is None:
        return None
    text = str(text).strip()
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def metric_for(row: dict[str, str]):
    task = value(row, "任务", "task")
    if task == "has_vul":
        return number(value(row, "f1"))
    if task == "vul_type":
        return number(value(row, "macro_f1", "f1"))
    if task == "vul_line":
        return number(value(row, "tolerant_f1", "strict_f1"))
    return None


def table_s7(rows: list[dict[str, str]]) -> list[dict[str, str]]:
    baseline_terms = ("codebert", "graphcodebert", "unixcoder", "codet5")
    return [r for r in rows if any(term in value(r, "模型", "model").lower() for term in baseline_terms)]


def table_s8(rows: list[dict[str, str]]) -> list[dict[str, str]]:
    return [r for r in rows if "qwen2.5-coder" in value(r, "模型", "model").lower()]


def table_s9(rows: list[dict[str, str]]) -> list[dict[str, str]]:
    main_models = ("qwen2.5-coder-7b-instruct", "qwen2.5-coder-32b-instruct")
    return [r for r in rows if value(r, "模型", "model").lower() in main_models]


def table_s11(rows: list[dict[str, str]]) -> list[dict[str, str]]:
    grouped: dict[tuple[str, str], dict[str, dict[str, str]]] = {}
    for row in rows:
        model = value(row, "模型", "model")
        task = value(row, "任务", "task")
        method = value(row, "方法", "method")
        if not model or not task or not method:
            continue
        grouped.setdefault((model, task), {})[method] = row

    comparisons = [
        ("prompt_vs_direct", "prompt", "direct"),
        ("qlora_vs_direct", "qlora", "direct"),
        ("full_vs_direct", "full", "direct"),
        ("full_vs_qlora", "full", "qlora"),
    ]
    out = []
    for (model, task), methods in grouped.items():
        for label, left, right in comparisons:
            if left not in methods or right not in methods:
                continue
            lv = metric_for(methods[left])
            rv = metric_for(methods[right])
            if lv is None or rv is None:
                continue
            out.append(
                {
                    "comparison": label,
                    "model": model,
                    "task": task,
                    "left_method": left,
                    "right_method": right,
                    "left_metric": f"{lv:.6f}",
                    "right_metric": f"{rv:.6f}",
                    "delta": f"{(lv - rv):.6f}",
                    "metric_used": "f1 for has_vul; macro_f1 for vul_type; tolerant_f1 for vul_line",
                }
            )
    return out


def table_s13() -> list[dict[str, str]]:
    pred_rows = read_csv(REPLICATION / "prediction_summaries" / "prediction_artifact_index.csv")
    prompt_rows = table_s9(read_csv(RESULTS / "prompt_ablation.csv"))
    out = []
    for row in pred_rows:
        out.append(
            {
                "source": "prediction_artifact_index",
                "run_name": value(row, "run_name"),
                "model": value(row, "model"),
                "task": value(row, "task"),
                "method": value(row, "method"),
                "prediction_rows": value(row, "prediction_rows"),
                "size_bytes": value(row, "size_bytes"),
                "sha256": value(row, "sha256"),
                "source_path": value(row, "source_path"),
            }
        )
    for row in prompt_rows:
        selected = {
            "source": "prompt_ablation_parse_summary",
            "run_name": value(row, "run_name", "运行名称"),
            "model": value(row, "模型"),
            "task": value(row, "任务"),
            "method": value(row, "模式"),
        }
        for key, val in row.items():
            key_l = key.lower()
            if any(token in key_l for token in ["parse", "empty", "format", "json", "fail"]):
                selected[key] = val
        out.append(selected)
    return out


def table_s14() -> list[dict[str, str]]:
    rows = []
    for base in [REPLICATION / "figures", REPLICATION / "figure_source"]:
        for path in sorted(base.rglob("*")):
            if path.is_file():
                rows.append(
                    {
                        "artifact_type": "figure" if path.suffix.lower() == ".svg" else "figure_source",
                        "path": path.relative_to(ROOT).as_posix(),
                        "size_bytes": str(path.stat().st_size),
                        "sha256": sha256(path),
                    }
                )
    return rows


def build_tables() -> list[dict[str, str]]:
    all_metrics = read_csv(RESULTS / "paper_experiment_metrics_all.csv")
    TABLES.mkdir(parents=True, exist_ok=True)
    produced = []
    for sid, title, source in TABLE_DEFS:
        if sid == "S7":
            rows = table_s7(all_metrics)
            sources = "replication/results/paper_experiment_metrics_all.csv"
        elif sid == "S8":
            rows = table_s8(all_metrics)
            sources = "replication/results/paper_experiment_metrics_all.csv"
        elif sid == "S9":
            rows = table_s9(read_csv(RESULTS / "prompt_ablation.csv"))
            sources = "replication/results/prompt_ablation.csv"
        elif sid == "S11":
            rows = table_s11(all_metrics)
            sources = "replication/results/paper_experiment_metrics_all.csv"
        elif sid == "S13":
            rows = table_s13()
            sources = "replication/prediction_summaries/prediction_artifact_index.csv; replication/results/prompt_ablation.csv"
        elif sid == "S14":
            rows = table_s14()
            sources = "replication/figures/; replication/figure_source/"
        else:
            rows = read_csv(RESULTS / source)
            sources = f"replication/results/{source}"

        rows = normalize_rows(rows)
        filename = f"Table_{sid}.csv"
        path = TABLES / filename
        write_csv(path, rows)
        table_id = f"Table {sid}"
        produced.append(
            {
                "supplementary_table": table_id,
                "title": title,
                "file": f"supplementary/tables/{filename}",
                "source": sources,
                "rows": str(len(rows)),
                "sha256": sha256(path),
                "key_columns": KEY_COLUMNS.get(table_id, ""),
            }
        )
    write_csv(SUPP / "Supplementary_Table_Index.csv", normalize_rows(produced))
    return produced


def clean_sheet_name(text: str) -> str:
    text = re.sub(r"[^A-Za-z0-9_]", "_", text)
    return text[:31]


def build_workbook(index: list[dict[str, str]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Index"
    write_sheet(ws, index)
    for item in index:
        csv_path = ROOT / item["file"]
        rows = read_csv(csv_path)
        ws = wb.create_sheet(clean_sheet_name(item["supplementary_table"].replace(" ", "_")))
        write_sheet(ws, rows)
    wb.save(SUPP / "supplementary_tables.xlsx")


def write_sheet(ws, rows: list[dict[str, str]], start_row: int = 1) -> None:
    fields = sorted({key for row in rows for key in row.keys()}) or ["note"]
    header_row = start_row
    for col, field in enumerate(fields, 1):
        cell = ws.cell(header_row, col, field)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="3B4A5A")
    for r_idx, row in enumerate(rows, header_row + 1):
        for c_idx, field in enumerate(fields, 1):
            ws.cell(r_idx, c_idx, row.get(field, ""))
    ws.freeze_panes = ws.cell(header_row + 1, 1).coordinate
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(fields))}{max(header_row + 1, header_row + len(rows))}"
    for col_idx, field in enumerate(fields, 1):
        max_len = len(field)
        for row in rows[:200]:
            max_len = max(max_len, len(str(row.get(field, ""))))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 48)


def write_readme(index: list[dict[str, str]]) -> None:
    lines = [
        "# Supplementary Tables",
        "",
        "This directory contains the numbered supplementary tables referenced by the manuscript. The table numbers are stable and match the in-text references in `Q1_English_IST_polished.md`.",
        "",
        "| Table | Title | CSV | Rows | Key columns | Source |",
        "|---|---|---|---:|---|---|",
    ]
    for item in index:
        lines.append(f"| {item['supplementary_table']} | {item['title']} | `{item['file']}` | {item['rows']} | {item.get('key_columns', '')} | {item['source']} |")
    lines.extend(
        [
            "",
            "Critical manuscript references:",
            "",
            "- Section 5 refers to Supplementary Table S1, available as `supplementary/tables/Table_S1.csv`.",
            "- Section 7.9 refers to Supplementary Tables S5, S9, and S13, available as `supplementary/tables/Table_S5.csv`, `supplementary/tables/Table_S9.csv`, and `supplementary/tables/Table_S13.csv`.",
            "",
            "A combined workbook is provided as `supplementary/supplementary_tables.xlsx` with worksheets named `Table_S1` to `Table_S14`. Each worksheet starts with the column header in row 1.",
        ]
    )
    (SUPP / "README.md").write_text("\n".join(lines) + "\n", encoding="utf-8")

    zh = [
        "# 补充材料表",
        "",
        "本目录存放论文正文引用的编号补充表。表号与 `Q1_English_IST_polished.md` 中的正文引用保持一致。",
        "",
        "| 表号 | 标题 | CSV | 行数 | 关键列 | 来源 |",
        "|---|---|---|---:|---|---|",
    ]
    for item in index:
        zh_title = ZH_TITLES.get(item["supplementary_table"], item["title"])
        zh.append(f"| {item['supplementary_table']} | {zh_title} | `{item['file']}` | {item['rows']} | {item.get('key_columns', '')} | {item['source']} |")
    zh.extend(
        [
            "",
            "重点正文引用核对：",
            "",
            "- 第 5 节引用 Supplementary Table S1，对应 `supplementary/tables/Table_S1.csv`。",
            "- 第 7.9 节引用 Supplementary Tables S5, S9, and S13，分别对应 `supplementary/tables/Table_S5.csv`、`supplementary/tables/Table_S9.csv` 和 `supplementary/tables/Table_S13.csv`。",
            "",
            "同时提供合并工作簿 `supplementary/supplementary_tables.xlsx`，其中工作表命名为 `Table_S1` 至 `Table_S14`；每个工作表第 1 行即为列名。",
        ]
    )
    (SUPP / "README.zh-CN.md").write_text("\n".join(zh) + "\n", encoding="utf-8")


def write_manifest() -> None:
    files = []
    for path in sorted(SUPP.rglob("*")):
        if path.is_file() and path.name != "manifest.json":
            files.append(
                {
                    "path": path.relative_to(ROOT).as_posix(),
                    "size_bytes": path.stat().st_size,
                    "sha256": sha256(path),
                }
            )
    (SUPP / "manifest.json").write_text(
        json.dumps({"file_count": len(files), "files": files}, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )


def main() -> None:
    index = build_tables()
    build_workbook(index)
    write_readme(index)
    write_manifest()


if __name__ == "__main__":
    main()
