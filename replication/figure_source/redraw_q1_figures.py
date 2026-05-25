from __future__ import annotations

import html
import math
import re
from collections import defaultdict
from pathlib import Path
from statistics import mean

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
PAPER_DIR = ROOT / "papers" / "当前论文"
XLSX_PATH = PAPER_DIR / "本地实验结果汇总_按模型任务结果.xlsx"
OUT_DIR = PAPER_DIR / "figures_q1_redesign"


PALETTE = {
    "direct": "#5A6E8C",
    "prompt": "#C9893F",
    "qlora": "#2A9D78",
    "full": "#C44E52",
    "baseline": "#7A7F87",
    "text": "#253041",
    "muted": "#6B7280",
    "grid": "#D7DDE8",
    "light": "#F5F7FA",
    "line": "#3D4656",
    "accent": "#3B82B8",
    "warn": "#B8563C",
}


def fnum(value):
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def fmt(value, digits=3):
    value = fnum(value)
    return "" if value is None else f"{value:.{digits}f}"


def esc(value):
    return html.escape(str(value), quote=True)


def read_sheet(wb, name):
    ws = wb[name]
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(x).strip() if x is not None else "" for x in rows[0]]
    out = []
    for row in rows[1:]:
        item = {}
        for header, value in zip(headers, row):
            item[header] = value
        out.append(item)
    return out


def short_model(name):
    mapping = {
        "Qwen/Qwen2.5-Coder-0.5B-Instruct": "Qwen-0.5B",
        "Qwen/Qwen2.5-Coder-1.5B-Instruct": "Qwen-1.5B",
        "Qwen/Qwen2.5-Coder-3B-Instruct": "Qwen-3B",
        "Qwen/Qwen2.5-Coder-7B-Instruct": "Qwen-7B",
        "Qwen/Qwen2.5-Coder-14B-Instruct": "Qwen-14B",
        "Qwen/Qwen2.5-Coder-32B-Instruct": "Qwen-32B",
        "deepseek-ai/deepseek-coder-1.3b-base": "DeepSeek-1.3B",
        "deepseek-ai/deepseek-coder-6.7b-base": "DeepSeek-6.7B",
        "deepseek-ai/deepseek-coder-33b-base": "DeepSeek-33B",
        "ibm-granite/granite-3b-code-base": "Granite-3B",
        "ibm-granite/granite-8b-code-base": "Granite-8B",
        "ibm-granite/granite-20b-code-base": "Granite-20B",
        "bigcode/starcoder2-3b": "StarCoder2-3B",
        "bigcode/starcoder2-7b": "StarCoder2-7B",
        "bigcode/starcoder2-15b": "StarCoder2-15B",
        "google/codegemma-2b": "CodeGemma-2B",
        "google/codegemma-7b": "CodeGemma-7B",
        "codellama/CodeLlama-7b-hf": "CodeLlama-7B",
        "codellama/CodeLlama-13b-hf": "CodeLlama-13B",
        "codellama/CodeLlama-34b-hf": "CodeLlama-34B",
        "WizardLM/WizardCoder-Python-7B-V1.0": "WizardCoder-7B",
        "microsoft/codebert-base": "CodeBERT",
        "microsoft/graphcodebert-base": "GraphCodeBERT",
        "microsoft/unixcoder-base": "UniXcoder",
        "Salesforce/codet5-base": "CodeT5",
    }
    return mapping.get(str(name), str(name).split("/")[-1])


def size_b(name):
    match = re.search(r"(\d+(?:\.\d+)?)B", str(name), re.IGNORECASE)
    return float(match.group(1)) if match else None


def text(x, y, value, size=10, weight=400, anchor="middle", color=None, rotate=None):
    color = color or PALETTE["text"]
    transform = f' transform="rotate({rotate} {x} {y})"' if rotate else ""
    return (
        f'<text x="{x:.1f}" y="{y:.1f}" font-family="Arial, Helvetica, sans-serif" '
        f'font-size="{size}" font-weight="{weight}" text-anchor="{anchor}" '
        f'fill="{color}"{transform}>{esc(value)}</text>'
    )


def wrap_text(x, y, value, width_chars=18, size=10, weight=400, anchor="middle", color=None, line_gap=13):
    words = str(value).replace("_", " ").split()
    lines = []
    current = ""
    for word in words:
        candidate = f"{current} {word}".strip()
        if len(candidate) > width_chars and current:
            lines.append(current)
            current = word
        else:
            current = candidate
    if current:
        lines.append(current)
    return [
        text(x, y + i * line_gap, line, size=size, weight=weight, anchor=anchor, color=color)
        for i, line in enumerate(lines[:3])
    ]


def svg(path, width, height, body):
    path.parent.mkdir(parents=True, exist_ok=True)
    doc = [
        f'<svg xmlns="http://www.w3.org/2000/svg" width="{width}" height="{height}" viewBox="0 0 {width} {height}">',
        '<rect width="100%" height="100%" fill="white"/>',
        *body,
        "</svg>",
    ]
    path.write_text("\n".join(doc), encoding="utf-8")


def axis_grid(body, left, top, plot_w, plot_h, xmax=1.0, ymax=1.0, x_ticks=5, y_ticks=5):
    for i in range(y_ticks + 1):
        y = top + plot_h - plot_h * i / y_ticks
        value = ymax * i / y_ticks
        body.append(f'<line x1="{left}" y1="{y:.1f}" x2="{left + plot_w}" y2="{y:.1f}" stroke="{PALETTE["grid"]}" stroke-width="1"/>')
        body.append(text(left - 10, y + 4, f"{value:.1f}", 9, anchor="end", color=PALETTE["muted"]))
    if x_ticks > 0:
        for i in range(x_ticks + 1):
            x = left + plot_w * i / x_ticks
            value = xmax * i / x_ticks
            body.append(f'<line x1="{x:.1f}" y1="{top}" x2="{x:.1f}" y2="{top + plot_h}" stroke="{PALETTE["grid"]}" stroke-width="1"/>')
            body.append(text(x, top + plot_h + 18, f"{value:.1f}", 9, color=PALETTE["muted"]))


def metric_for(row):
    task = row.get("任务")
    if task == "has_vul":
        return fnum(row.get("f1"))
    if task == "vul_type":
        return fnum(row.get("macro_f1"))
    if task == "vul_line":
        return fnum(row.get("tolerant_f1"))
    return None


def formal_rows(rows):
    seen = set()
    clean = []
    for row in rows:
        run_name = str(row.get("run_name") or "")
        if run_name in seen:
            continue
        seen.add(run_name)
        if row.get("系列") == "补齐前后对比":
            continue
        clean.append(row)
    return clean


def draw_fig1_protocol(rows):
    width, height = 1100, 520
    body = [text(width / 2, 34, "Unified protocol for multitask LLM vulnerability analysis", 18, 700)]
    body.append('<defs><marker id="arrow" markerWidth="9" markerHeight="9" refX="8" refY="3" orient="auto"><path d="M0,0 L0,6 L8,3 z" fill="#68758A"/></marker></defs>')
    cards = [
        ("Data governance", "multi-source code\nlabel mapping\ndeduplicated splits"),
        ("Tasks", "has_vul\nvul_type\nvul_line"),
        ("Modes", "direct\nprompt\nfull fine-tune\nQLoRA"),
        ("Models", "open LLMs\ncode baselines\nfamily/scale slices"),
        ("Metrics", "F1/macro-F1\nstrict/tolerant F1\ncontract hit"),
        ("Evidence synthesis", "capability boundary\ncost/stability\nfailure modes"),
    ]
    x0, y0, w, h, gap = 50, 105, 148, 160, 31
    for i, (title, desc) in enumerate(cards):
        x = x0 + i * (w + gap)
        body.append(f'<rect x="{x}" y="{y0}" width="{w}" height="{h}" rx="7" fill="{PALETTE["light"]}" stroke="#C6CDD8"/>')
        body.append(text(x + w / 2, y0 + 28, title, 12, 700))
        for line_i, line in enumerate(desc.split("\n")):
            body.append(text(x + w / 2, y0 + 66 + line_i * 22, line, 11, 400, color=PALETTE["muted"]))
        if i < len(cards) - 1:
            body.append(f'<path d="M{x + w + 7},{y0 + h / 2} L{x + w + gap - 9},{y0 + h / 2}" stroke="#68758A" stroke-width="1.8" marker-end="url(#arrow)"/>')
    body.append(f'<rect x="90" y="330" width="920" height="118" rx="7" fill="#FFFFFF" stroke="#B9C2D0"/>')
    body.append(text(550, 358, "Closed formal matrix: 226 planned experiments, 226 completed, 0 missing", 14, 700))
    body.append(text(550, 388, "Main figures use one weight per unique run_name; data-completion pairs are analyzed separately.", 12, color=PALETTE["muted"]))
    body.append(text(550, 418, "The study is framed as an empirical software-engineering benchmark rather than a single-model leaderboard.", 12, color=PALETTE["muted"]))
    svg(OUT_DIR / "fig1_protocol.svg", width, height, body)


def draw_fig2_task_hierarchy(rows):
    width, height = 980, 600
    body = [text(width / 2, 34, "Task hierarchy across best and average model performance", 18, 700)]
    tasks = [
        ("has_vul", "Has-vul F1", "f1"),
        ("vul_type", "Type macro-F1", "macro_f1"),
        ("vul_line", "Line strict-F1", "strict_f1"),
        ("vul_line", "Line tolerant-F1", "tolerant_f1"),
        ("vul_line", "Line contract-hit", "contract_hit"),
    ]
    labels, best_vals, avg_vals = [], [], []
    for task, label, metric in tasks:
        vals = [fnum(r.get(metric)) for r in rows if r.get("任务") == task and fnum(r.get(metric)) is not None]
        labels.append(label)
        best_vals.append(max(vals))
        avg_vals.append(mean(vals))
    left, top, plot_w, plot_h = 88, 78, 835, 390
    axis_grid(body, left, top, plot_w, plot_h, y_ticks=5)
    group_w = plot_w / len(labels)
    for i, label in enumerate(labels):
        cx = left + group_w * i + group_w / 2
        for off, value, color, name in [(-19, avg_vals[i], "#9AA6B8", "mean"), (19, best_vals[i], PALETTE["accent"], "best")]:
            h = plot_h * value
            body.append(f'<rect x="{cx + off - 13:.1f}" y="{top + plot_h - h:.1f}" width="26" height="{h:.1f}" rx="2" fill="{color}"/>')
            body.append(text(cx + off, top + plot_h - h - 6, fmt(value), 9, color=color))
        body.extend(wrap_text(cx, height - 86, label, width_chars=14, size=10))
    for i, (name, color) in enumerate([("mean", "#9AA6B8"), ("best", PALETTE["accent"])]):
        x = 760 + i * 82
        body.append(f'<rect x="{x}" y="520" width="14" height="14" fill="{color}"/>')
        body.append(text(x + 20, 531, name, 11, anchor="start"))
    body.append(text(left + plot_w / 2, 568, "Scores are task-specific metrics; higher is better.", 11, color=PALETTE["muted"]))
    svg(OUT_DIR / "fig2_task_hierarchy.svg", width, height, body)


def draw_fig3_method_deltas(rows):
    width, height = 1060, 610
    body = [text(width / 2, 34, "Paired method gains over the same model and task", 18, 700)]
    comparisons = [
        ("prompt-direct", "prompt", "direct"),
        ("qlora-direct", "qlora", "direct"),
        ("full-direct", "full", "direct"),
        ("full-qlora", "full", "qlora"),
    ]
    tasks = [("has_vul", "Has-F1", "f1"), ("vul_type", "Type macro-F1", "macro_f1"), ("vul_line", "Line tolerant-F1", "tolerant_f1")]
    left, top, plot_w, plot_h = 92, 84, 884, 380
    min_v, max_v = -0.18, 0.86
    zero_x = left + plot_w * (0 - min_v) / (max_v - min_v)
    for i in range(7):
        value = min_v + (max_v - min_v) * i / 6
        x = left + plot_w * (value - min_v) / (max_v - min_v)
        body.append(f'<line x1="{x:.1f}" y1="{top}" x2="{x:.1f}" y2="{top + plot_h}" stroke="{PALETTE["grid"]}" stroke-width="1"/>')
        body.append(text(x, top + plot_h + 18, f"{value:.1f}", 9, color=PALETTE["muted"]))
    body.append(f'<line x1="{zero_x:.1f}" y1="{top - 4}" x2="{zero_x:.1f}" y2="{top + plot_h + 4}" stroke="#1F2937" stroke-width="1.4"/>')
    row_gap = plot_h / (len(comparisons) * len(tasks) + 1)
    y = top + row_gap
    for ci, (comp, high, low) in enumerate(comparisons):
        body.append(text(28, y + row_gap, comp, 11, 700, anchor="start"))
        for task, label, metric in tasks:
            pivot = defaultdict(dict)
            for row in rows:
                if row.get("任务") == task:
                    pivot[row.get("模型")][row.get("方法")] = fnum(row.get(metric))
            deltas = [v[high] - v[low] for v in pivot.values() if v.get(high) is not None and v.get(low) is not None]
            if not deltas:
                continue
            avg, lo, hi = mean(deltas), min(deltas), max(deltas)
            x0 = left + plot_w * (lo - min_v) / (max_v - min_v)
            x1 = left + plot_w * (hi - min_v) / (max_v - min_v)
            xa = left + plot_w * (avg - min_v) / (max_v - min_v)
            color = PALETTE[high]
            body.append(f'<line x1="{x0:.1f}" y1="{y:.1f}" x2="{x1:.1f}" y2="{y:.1f}" stroke="{color}" stroke-width="2.2" opacity="0.65"/>')
            body.append(f'<circle cx="{xa:.1f}" cy="{y:.1f}" r="5.4" fill="{color}" stroke="white" stroke-width="1"/>')
            body.append(text(left - 12, y + 4, label, 10, anchor="end"))
            body.append(text(x1 + 8, y + 4, f"n={len(deltas)}, mean={avg:.3f}", 9, anchor="start", color=PALETTE["muted"]))
            y += row_gap
        y += row_gap * 0.5
    body.append(text(left + plot_w / 2, height - 38, "Delta in task-specific score; whisker shows min-max across paired models.", 11, color=PALETTE["muted"]))
    svg(OUT_DIR / "fig3_method_deltas.svg", width, height, body)


def draw_fig4_qwen_scaling(rows):
    width, height = 1120, 650
    body = [text(width / 2, 34, "Qwen2.5-Coder scaling is non-monotonic across modes", 18, 700)]
    qwen = [r for r in rows if str(r.get("系列")) == "Qwen2.5-Coder"]
    panels = [("has_vul", "Has-vul F1", "f1"), ("vul_type", "Type macro-F1", "macro_f1"), ("vul_line", "Line tolerant-F1", "tolerant_f1")]
    sizes = sorted({size_b(r.get("模型")) for r in qwen if size_b(r.get("模型")) is not None})
    methods = ["direct", "prompt", "qlora", "full"]
    top = 88
    panel_w = 330
    plot_h = 390
    for pi, (task, title, metric) in enumerate(panels):
        left = 64 + pi * (panel_w + 28)
        body.append(text(left + panel_w / 2, 66, title, 13, 700))
        axis_grid(body, left, top, panel_w, plot_h, x_ticks=0, y_ticks=5)
        log_min = math.log(min(sizes) + 0.2)
        log_max = math.log(max(sizes) + 0.2)

        def sx(value):
            return left + (math.log(value + 0.2) - log_min) / (log_max - log_min) * panel_w

        def sy(value):
            return top + plot_h - max(0, min(value, 1)) * plot_h

        for method in methods:
            pts = []
            for size in sizes:
                candidates = [
                    r for r in qwen
                    if r.get("任务") == task and r.get("方法") == method and size_b(r.get("模型")) == size
                ]
                if candidates and fnum(candidates[0].get(metric)) is not None:
                    pts.append((size, fnum(candidates[0].get(metric))))
            if not pts:
                continue
            d = []
            for i, (size, value) in enumerate(pts):
                d.append(("M" if i == 0 else "L") + f" {sx(size):.1f} {sy(value):.1f}")
            body.append(f'<path d="{" ".join(d)}" fill="none" stroke="{PALETTE[method]}" stroke-width="2.2"/>')
            for size, value in pts:
                body.append(f'<circle cx="{sx(size):.1f}" cy="{sy(value):.1f}" r="4.5" fill="{PALETTE[method]}" stroke="white" stroke-width="1"/>')
        for size in sizes:
            body.append(text(sx(size), top + plot_h + 24, f"{size:g}B", 9, color=PALETTE["muted"], rotate=-35))
    legend_y = height - 52
    for i, method in enumerate(methods):
        x = 380 + i * 96
        body.append(f'<line x1="{x}" y1="{legend_y}" x2="{x + 24}" y2="{legend_y}" stroke="{PALETTE[method]}" stroke-width="3"/>')
        body.append(text(x + 31, legend_y + 4, method, 11, anchor="start"))
    svg(OUT_DIR / "fig4_qwen_scaling.svg", width, height, body)


def draw_fig5_prompt(prompt):
    width, height = 1120, 630
    body = [text(width / 2, 34, "Prompt ablation: task scores and output failures diverge", 18, 700)]
    prompt = [r for r in prompt if "7B" in str(r.get("模型")) or "32B" in str(r.get("模型"))]
    panels = [
        ("vul_type", "Type macro-F1", "macro_f1"),
        ("vul_line", "Line tolerant-F1", "tolerant_f1"),
        ("vul_line", "Empty output rate", "empty_output_rate"),
    ]
    models = ["Qwen2.5-Coder-7B-Instruct", "Qwen2.5-Coder-32B-Instruct"]
    versions = ["P0", "P1", "P2", "P3"]
    top = 88
    panel_w = 310
    plot_h = 360
    for pi, (task, title, metric) in enumerate(panels):
        left = 74 + pi * (panel_w + 44)
        body.append(text(left + panel_w / 2, 66, title, 13, 700))
        axis_grid(body, left, top, panel_w, plot_h, x_ticks=0, y_ticks=5)
        group_w = panel_w / len(models)
        bar_w = 17
        for mi, model in enumerate(models):
            cx = left + group_w * mi + group_w / 2
            for vi, version in enumerate(versions):
                candidates = [
                    r for r in prompt
                    if str(r.get("模型")) == model and r.get("任务") == task and r.get("版本") == version
                ]
                value = fnum(candidates[0].get(metric)) if candidates else 0
                x = cx - 2 * bar_w + vi * bar_w
                h = plot_h * max(0, min(value, 1))
                body.append(f'<rect x="{x:.1f}" y="{top + plot_h - h:.1f}" width="{bar_w - 3}" height="{h:.1f}" fill="{["#64748B", "#C9893F", "#2A9D78", "#C44E52"][vi]}"/>')
                body.append(text(x + 7, top + plot_h - h - 5, fmt(value, 2), 8, color=PALETTE["muted"]))
            body.append(text(cx, top + plot_h + 30, model.replace("Qwen2.5-Coder-", "").replace("-Instruct", ""), 10, color=PALETTE["muted"]))
    for vi, version in enumerate(versions):
        x = 398 + vi * 82
        body.append(f'<rect x="{x}" y="570" width="14" height="14" fill="{["#64748B", "#C9893F", "#2A9D78", "#C44E52"][vi]}"/>')
        body.append(text(x + 20, 581, version, 11, anchor="start"))
    svg(OUT_DIR / "fig5_prompt_ablation.svg", width, height, body)


def draw_fig6_completion(completion_rows):
    width, height = 940, 590
    body = [text(width / 2, 34, "Dataset completion changes measured difficulty", 18, 700)]
    pairs = [
        ("Qwen-3B type", "Qwen/Qwen2.5-Coder-3B-Instruct", "vul_type", "qlora", "f1"),
        ("Qwen-3B line", "Qwen/Qwen2.5-Coder-3B-Instruct", "vul_line", "qlora", "strict_f1"),
        ("CodeBERT type", "microsoft/codebert-base", "vul_type", "full", "f1"),
        ("CodeBERT line", "microsoft/codebert-base", "vul_line", "full", "strict_f1"),
        ("GraphCodeBERT type", "microsoft/graphcodebert-base", "vul_type", "full", "f1"),
        ("GraphCodeBERT line", "microsoft/graphcodebert-base", "vul_line", "full", "strict_f1"),
    ]
    left, right = 330, 715
    top, gap = 102, 67
    body.append(text(left, 73, "Original 3000-sample stage", 12, 700))
    body.append(text(right, 73, "Completed large-scale stage", 12, 700))
    for i, (label, model, task, method, metric) in enumerate(pairs):
        before = [
            r for r in completion_rows
            if r.get("模型") == model and r.get("任务") == task and r.get("方法") == method and r.get("补齐阶段说明") == "补齐前"
        ]
        after = [
            r for r in completion_rows
            if r.get("模型") == model and r.get("任务") == task and r.get("方法") == method and r.get("补齐阶段说明") == "补齐后"
        ]
        if not before or not after:
            continue
        b, a = fnum(before[0].get(metric)), fnum(after[0].get(metric))
        y = top + i * gap
        color = PALETTE["accent"] if task == "vul_type" else PALETTE["warn"]
        body.append(text(80, y + 4, label, 11, 700, anchor="start"))
        body.append(text(230, y + 4, metric, 10, anchor="start", color=PALETTE["muted"]))
        body.append(f'<line x1="{left}" y1="{y}" x2="{right}" y2="{y}" stroke="{color}" stroke-width="2.2" opacity="0.8"/>')
        body.append(f'<circle cx="{left}" cy="{y}" r="6" fill="{color}"/>')
        body.append(f'<circle cx="{right}" cy="{y}" r="6" fill="{color}"/>')
        body.append(text(left - 12, y + 4, fmt(b), 10, anchor="end"))
        body.append(text(right + 12, y + 4, fmt(a), 10, anchor="start"))
        body.append(text((left + right) / 2, y - 8, f"{a - b:+.3f}", 10, 700, color=color))
    body.append(text(width / 2, height - 34, "Type and line tasks are shown separately because completion changes label coverage and localization difficulty.", 11, color=PALETTE["muted"]))
    svg(OUT_DIR / "fig6_data_completion.svg", width, height, body)


def draw_fig8_line_boundary(rows):
    width, height = 930, 650
    body = [text(width / 2, 34, "Vulnerable-line localization separates candidate hits from exact lines", 18, 700)]
    left, top, plot_w, plot_h = 94, 80, 750, 450
    axis_grid(body, left, top, plot_w, plot_h, xmax=0.35, ymax=0.9, x_ticks=7, y_ticks=6)
    body.append(text(left + plot_w / 2, height - 57, "Strict line F1", 12, 700))
    body.append(text(22, top + plot_h / 2, "Contract-hit", 12, 700))
    for row in rows:
        if row.get("任务") != "vul_line":
            continue
        strict = fnum(row.get("strict_f1"))
        hit = fnum(row.get("contract_hit"))
        if strict is None or hit is None:
            continue
        x = left + plot_w * strict / 0.35
        y = top + plot_h - plot_h * hit / 0.9
        radius = 4.2 + min(math.log((size_b(row.get("模型")) or 1) + 1) * 1.8, 6)
        method = row.get("方法")
        body.append(f'<circle cx="{x:.1f}" cy="{y:.1f}" r="{radius:.1f}" fill="{PALETTE.get(method, PALETTE["baseline"])}" fill-opacity="0.72" stroke="white" stroke-width="1"/>')
    body.append(f'<rect x="{left + 500}" y="{top + 40}" width="190" height="78" rx="6" fill="white" stroke="#C9D2E3"/>')
    body.append(text(left + 595, top + 65, "high hit, low strict-F1", 11, 700))
    body.append(text(left + 595, top + 91, "candidate regions are useful", 10, color=PALETTE["muted"]))
    body.append(text(left + 595, top + 109, "exact line numbers remain hard", 10, color=PALETTE["muted"]))
    for i, method in enumerate(["direct", "prompt", "qlora", "full"]):
        x = 600 + i * 76
        body.append(f'<circle cx="{x}" cy="598" r="6" fill="{PALETTE[method]}"/>')
        body.append(text(x + 10, 602, method, 10, anchor="start"))
    svg(OUT_DIR / "fig8_line_boundary.svg", width, height, body)


def draw_fig7_family_heatmap(rows):
    width, height = 1080, 720
    body = [text(width / 2, 34, "Model family, task, and adaptation mode jointly shape performance", 18, 700)]
    families = [x for x in sorted({r.get("系列") for r in rows if r.get("系列")}) if x != "补齐前后对比"]
    tasks = [("has_vul", "Has-F1"), ("vul_type", "Type macro-F1"), ("vul_line", "Line tolerant-F1")]
    methods = ["direct", "prompt", "qlora", "full"]
    left, top = 205, 84
    cell_w, cell_h = 60, 32
    gap_task = 20
    for ti, (_, task_label) in enumerate(tasks):
        x0 = left + ti * (len(methods) * cell_w + gap_task)
        body.append(text(x0 + len(methods) * cell_w / 2, 65, task_label, 12, 700))
        for mi, method in enumerate(methods):
            body.append(text(x0 + mi * cell_w + cell_w / 2, top - 8, method[0].upper(), 10, 700, color=PALETTE[method]))
    for fi, family in enumerate(families):
        y = top + fi * cell_h
        body.append(text(left - 14, y + 20, family, 10, anchor="end"))
        for ti, (task, _) in enumerate(tasks):
            x0 = left + ti * (len(methods) * cell_w + gap_task)
            for mi, method in enumerate(methods):
                vals = [
                    metric_for(r) for r in rows
                    if r.get("系列") == family and r.get("任务") == task and r.get("方法") == method and metric_for(r) is not None
                ]
                val = mean(vals) if vals else None
                if val is None:
                    fill, label = "#EEF1F6", "-"
                else:
                    t = max(0, min(val / 0.88, 1))
                    r = int(245 - 120 * t)
                    g = int(247 - 52 * t)
                    b = int(250 - 95 * t)
                    fill, label = f"rgb({r},{g},{b})", f"{val:.2f}"
                x = x0 + mi * cell_w
                body.append(f'<rect x="{x}" y="{y}" width="{cell_w - 3}" height="{cell_h - 3}" fill="{fill}" stroke="white"/>')
                body.append(text(x + cell_w / 2, y + 20, label, 9))
    body.append(text(width / 2, height - 38, "D/P/Q/F denote direct, prompt, QLoRA, and full fine-tuning.", 11, color=PALETTE["muted"]))
    svg(OUT_DIR / "fig7_family_task_mode_heatmap.svg", width, height, body)


def draw_fig9_runtime(rows):
    width, height = 1000, 620
    body = [text(width / 2, 34, "Performance-cost trade-offs expose deployment-relevant choices", 18, 700)]
    # Runtime exists on the broader all-results sheet, so rows passed here may include runtime_minutes.
    task_specs = [("has_vul", "Has-F1", "f1"), ("vul_type", "Type macro-F1", "macro_f1"), ("vul_line", "Line tolerant-F1", "tolerant_f1")]
    panels = []
    for task, label, metric in task_specs:
        pts = []
        for row in rows:
            if row.get("任务") != task:
                continue
            score = fnum(row.get(metric))
            runtime = fnum(row.get("runtime_minutes"))
            if score is None or runtime is None or runtime <= 0:
                continue
            pts.append((runtime, score, row))
        panels.append((label, pts))
    top = 90
    panel_w = 286
    plot_h = 370
    for pi, (label, pts) in enumerate(panels):
        left = 64 + pi * (panel_w + 38)
        body.append(text(left + panel_w / 2, 66, label, 13, 700))
        max_runtime = max([p[0] for p in pts] or [1])
        x_max = max(100, max_runtime)
        for i in range(6):
            y = top + plot_h - plot_h * i / 5
            body.append(f'<line x1="{left}" y1="{y:.1f}" x2="{left + panel_w}" y2="{y:.1f}" stroke="{PALETTE["grid"]}" stroke-width="1"/>')
            body.append(text(left - 8, y + 4, f"{i/5:.1f}", 8, anchor="end", color=PALETTE["muted"]))
        for runtime, score, row in pts:
            x = left + math.log(runtime + 1) / math.log(x_max + 1) * panel_w
            y = top + plot_h - max(0, min(score, 1)) * plot_h
            body.append(f'<circle cx="{x:.1f}" cy="{y:.1f}" r="4.8" fill="{PALETTE.get(row.get("方法"), PALETTE["baseline"])}" fill-opacity="0.7" stroke="white" stroke-width="0.8"/>')
        body.append(text(left + panel_w / 2, top + plot_h + 28, "runtime minutes (log scale)", 9, color=PALETTE["muted"]))
    for i, method in enumerate(["direct", "prompt", "qlora", "full"]):
        x = 356 + i * 90
        body.append(f'<circle cx="{x}" cy="574" r="6" fill="{PALETTE[method]}"/>')
        body.append(text(x + 11, 578, method, 10, anchor="start"))
    svg(OUT_DIR / "fig9_runtime_tradeoff.svg", width, height, body)


def write_figure_notes(rows, formal_rows_raw, all_rows, prompt, completion):
    path = OUT_DIR / "figure_redesign_notes.md"
    excluded_completion_rows = [r for r in formal_rows_raw if r.get("系列") == "补齐前后对比"]
    duplicated_run_rows = [
        r for r in formal_rows_raw
        if str(r.get("run_name") or "") and sum(1 for x in formal_rows_raw if x.get("run_name") == r.get("run_name")) > 1
    ]
    lines = [
        "# Q1 figure redesign notes",
        "",
        "Source workbook: `本地实验结果汇总_按模型任务结果.xlsx`.",
        "",
        "## Statistical scope",
        "",
        f"- Main formal figures use {len(rows)} records after excluding the narrative-only `补齐前后对比` rows.",
        "- The closure statement remains `226/226/0` because the workbook contains 226 completed formal rows.",
        "- Fig.6 uses the six explicit before/after pairs from `补齐前后12条明细`.",
        f"- Excluded `补齐前后对比` rows from main statistics: {len(excluded_completion_rows)}.",
        f"- Rows whose `run_name` is duplicated in the raw formal table: {len(duplicated_run_rows)}.",
        "",
        "## Figure set",
        "",
        "- Fig.1: protocol and closed experiment matrix.",
        "- Fig.2: task hierarchy using best and mean task-specific scores.",
        "- Fig.3: paired method deltas within the same model and task.",
        "- Fig.4: Qwen2.5-Coder scaling by mode.",
        "- Fig.5: prompt ablation for 7B/32B, including output failure rates.",
        "- Fig.6: dataset completion before/after pairs.",
        "- Fig.7: family-task-mode heatmap.",
        "- Fig.8: line localization exactness versus contract-level hit.",
        "- Fig.9: runtime/performance trade-off.",
    ]
    path.write_text("\n".join(lines), encoding="utf-8")


def main():
    wb = load_workbook(XLSX_PATH, data_only=True)
    formal = read_sheet(wb, "论文实验指标总表")
    all_results = read_sheet(wb, "全部结果指标")
    prompt = read_sheet(wb, "Prompt消融")
    completion = read_sheet(wb, "补齐前后12条明细")

    rows = formal_rows(formal)
    all_clean = formal_rows(all_results)

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    draw_fig1_protocol(rows)
    draw_fig2_task_hierarchy(rows)
    draw_fig3_method_deltas(rows)
    draw_fig4_qwen_scaling(rows)
    draw_fig5_prompt(prompt)
    draw_fig6_completion(completion)
    draw_fig7_family_heatmap(rows)
    draw_fig8_line_boundary(rows)
    draw_fig9_runtime(all_clean)
    write_figure_notes(rows, formal, all_results, prompt, completion)
    print(f"Wrote redesigned figures to: {OUT_DIR}")


if __name__ == "__main__":
    main()
