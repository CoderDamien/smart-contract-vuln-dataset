from __future__ import annotations

import html
import math
import re
import zipfile
import xml.etree.ElementTree as ET
from collections import Counter, defaultdict
from pathlib import Path
from statistics import mean, median

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
PAPER_DIR = ROOT / "papers" / "当前论文"
XLSX_PATH = PAPER_DIR / "本地实验结果汇总_按模型任务结果.xlsx"
FIG_DIR = PAPER_DIR / "figures_q1"
DRAFT_PATH = PAPER_DIR / "Q1_中文完整草稿.md"
SUMMARY_PATH = PAPER_DIR / "Q1_数据表与图表摘要.md"
BOOKTABS_PATH = PAPER_DIR / "Q1_三线表_booktabs.tex"
HIGHLIGHTS_PATH = PAPER_DIR / "IST_highlights.md"


COLORS = {
    "direct": "#5B6C8F",
    "prompt": "#D98C40",
    "qlora": "#2E9F7D",
    "full": "#C94545",
    "baseline": "#8E8E8E",
    "grid": "#D9DEE8",
    "text": "#243044",
    "light": "#F6F8FB",
}


def fnum(value):
    if value is None:
        return None
    s = str(value).strip()
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def fmt(value, digits=4):
    v = fnum(value)
    if v is None:
        return ""
    return f"{v:.{digits}f}"


def esc(text):
    return html.escape(str(text), quote=True)


def short_model(name: str) -> str:
    mapping = {
        "Qwen/Qwen2.5-Coder-0.5B-Instruct": "Qwen-0.5B",
        "Qwen/Qwen2.5-Coder-1.5B-Instruct": "Qwen-1.5B",
        "Qwen/Qwen2.5-Coder-3B-Instruct": "Qwen-3B",
        "Qwen/Qwen2.5-Coder-7B-Instruct": "Qwen-7B",
        "Qwen/Qwen2.5-Coder-14B-Instruct": "Qwen-14B",
        "Qwen/Qwen2.5-Coder-32B-Instruct": "Qwen-32B",
        "ibm-granite/granite-3b-code-base": "Granite-3B",
        "ibm-granite/granite-8b-code-base": "Granite-8B",
        "ibm-granite/granite-20b-code-base": "Granite-20B",
        "deepseek-ai/deepseek-coder-1.3b-base": "DeepSeek-1.3B",
        "deepseek-ai/deepseek-coder-6.7b-base": "DeepSeek-6.7B",
        "deepseek-ai/deepseek-coder-33b-base": "DeepSeek-33B",
        "bigcode/starcoder2-3b": "StarCoder2-3B",
        "bigcode/starcoder2-7b": "StarCoder2-7B",
        "bigcode/starcoder2-15b": "StarCoder2-15B",
        "google/codegemma-2b": "CodeGemma-2B",
        "google/codegemma-7b": "CodeGemma-7B",
        "codellama/CodeLlama-7b-hf": "CodeLlama-7B",
        "codellama/CodeLlama-13b-hf": "CodeLlama-13B",
        "codellama/CodeLlama-34b-hf": "CodeLlama-34B",
        "WizardLM/WizardCoder-Python-7B-V1.0": "WizardCoder-7B",
        "microsoft/codebert-base": "CodeBERT",
        "microsoft/graphcodebert-base": "GraphCodeBERT",
        "microsoft/unixcoder-base": "UniXcoder",
        "Salesforce/codet5-base": "CodeT5",
    }
    return mapping.get(name, name.split("/")[-1])


def param_size(name: str):
    m = re.search(r"(\d+(?:\.\d+)?)B", name, re.IGNORECASE)
    return float(m.group(1)) if m else None


def read_sheet(wb, name: str):
    ws = wb[name]
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(x).strip() if x is not None else "" for x in rows[0]]
    out = []
    for row in rows[1:]:
        item = {}
        for h, v in zip(headers, row):
            item[h] = "" if v is None else v
        out.append(item)
    return out


def md_table(headers, rows):
    lines = []
    lines.append("| " + " | ".join(headers) + " |")
    lines.append("| " + " | ".join(["---" for _ in headers]) + " |")
    for row in rows:
        lines.append("| " + " | ".join(str(x) for x in row) + " |")
    return "\n".join(lines)


def tex_escape(text):
    s = str(text)
    for old, new in [
        ("\\", r"\textbackslash{}"),
        ("&", r"\&"),
        ("%", r"\%"),
        ("$", r"\$"),
        ("#", r"\#"),
        ("_", r"\_"),
        ("{", r"\{"),
        ("}", r"\}"),
        ("~", r"\textasciitilde{}"),
        ("^", r"\textasciicircum{}"),
    ]:
        s = s.replace(old, new)
    return s


def latex_table(caption, label, headers, rows, align=None):
    align = align or ("l" * len(headers))
    lines = [
        r"\begin{table}[htbp]",
        r"\centering",
        rf"\caption{{{tex_escape(caption)}}}",
        rf"\label{{{label}}}",
        rf"\begin{{tabular}}{{{align}}}",
        r"\toprule",
        " & ".join(tex_escape(h) for h in headers) + r" \\",
        r"\midrule",
    ]
    for row in rows:
        lines.append(" & ".join(tex_escape(x) for x in row) + r" \\")
    lines.extend([r"\bottomrule", r"\end{tabular}", r"\end{table}", ""])
    return "\n".join(lines)


def svg_text(x, y, text, size=12, weight="400", anchor="middle", color=None):
    return (
        f'<text x="{x}" y="{y}" font-family="Aptos, Calibri, Arial, sans-serif" '
        f'font-size="{size}" font-weight="{weight}" text-anchor="{anchor}" '
        f'fill="{color or COLORS["text"]}">{esc(text)}</text>'
    )


def save_svg(path: Path, width: int, height: int, body: list[str]):
    path.parent.mkdir(parents=True, exist_ok=True)
    svg = [
        f'<svg xmlns="http://www.w3.org/2000/svg" width="{width}" height="{height}" viewBox="0 0 {width} {height}">',
        f'<rect width="{width}" height="{height}" rx="18" fill="white"/>',
        *body,
        "</svg>",
    ]
    path.write_text("\n".join(svg), encoding="utf-8")


def draw_grouped_bar(path, labels, series, title, ylabel="Score"):
    width, height = 980, 560
    margin = dict(left=88, right=36, top=70, bottom=92)
    plot_w = width - margin["left"] - margin["right"]
    plot_h = height - margin["top"] - margin["bottom"]
    ymax = 1.0
    body = [
        svg_text(width / 2, 36, title, 20, "700"),
        svg_text(20, margin["top"] + plot_h / 2, ylabel, 12, anchor="middle"),
    ]
    for i in range(6):
        y = margin["top"] + plot_h - plot_h * i / 5
        val = ymax * i / 5
        body.append(f'<line x1="{margin["left"]}" y1="{y:.1f}" x2="{width-margin["right"]}" y2="{y:.1f}" stroke="{COLORS["grid"]}" stroke-width="1"/>')
        body.append(svg_text(margin["left"] - 14, y + 4, f"{val:.1f}", 11, anchor="end", color="#667085"))
    group_w = plot_w / len(labels)
    bar_w = min(30, group_w / (len(series) + 1.1))
    for si, (name, values, color) in enumerate(series):
        for i, v in enumerate(values):
            x = margin["left"] + group_w * i + group_w / 2 - (len(series) * bar_w) / 2 + si * bar_w
            h = plot_h * max(0, min(v, ymax)) / ymax
            y = margin["top"] + plot_h - h
            body.append(f'<rect x="{x:.1f}" y="{y:.1f}" width="{bar_w-3:.1f}" height="{h:.1f}" rx="4" fill="{color}"/>')
    for i, label in enumerate(labels):
        x = margin["left"] + group_w * i + group_w / 2
        body.append(svg_text(x, height - 54, label, 12))
    legend_x = margin["left"]
    for si, (name, _, color) in enumerate(series):
        x = legend_x + si * 130
        body.append(f'<rect x="{x}" y="{height-28}" width="14" height="14" rx="3" fill="{color}"/>')
        body.append(svg_text(x + 20, height - 16, name, 12, anchor="start"))
    save_svg(path, width, height, body)


def draw_line_panels(path, panels, title):
    width, height = 1120, 660
    body = [svg_text(width / 2, 36, title, 20, "700")]
    panel_w = (width - 90) / len(panels)
    top, bottom = 82, 96
    plot_h = height - top - bottom
    for pi, panel in enumerate(panels):
        left = 55 + pi * panel_w
        right = left + panel_w - 34
        body.append(svg_text((left + right) / 2, 66, panel["title"], 14, "700"))
        for j in range(6):
            y = top + plot_h - plot_h * j / 5
            body.append(f'<line x1="{left}" y1="{y:.1f}" x2="{right}" y2="{y:.1f}" stroke="{COLORS["grid"]}" stroke-width="1"/>')
            if pi == 0:
                body.append(svg_text(left - 12, y + 4, f"{j/5:.1f}", 10, anchor="end", color="#667085"))
        xs = panel["x"]
        if not xs:
            continue
        min_x, max_x = min(xs), max(xs)
        def sx(x):
            if max_x == min_x:
                return (left + right) / 2
            return left + (math.log(x + 0.2) - math.log(min_x + 0.2)) / (math.log(max_x + 0.2) - math.log(min_x + 0.2)) * (right - left)
        def sy(y):
            return top + plot_h - max(0, min(y, 1)) * plot_h
        for method, values in panel["series"].items():
            pts = [(x, v) for x, v in values if v is not None]
            if len(pts) < 1:
                continue
            d = " ".join(("M" if i == 0 else "L") + f" {sx(x):.1f} {sy(v):.1f}" for i, (x, v) in enumerate(pts))
            body.append(f'<path d="{d}" fill="none" stroke="{COLORS.get(method, "#333")}" stroke-width="2.6"/>')
            for x, v in pts:
                body.append(f'<circle cx="{sx(x):.1f}" cy="{sy(v):.1f}" r="4.5" fill="{COLORS.get(method, "#333")}"/>')
        for x in xs:
            body.append(svg_text(sx(x), height - 58, str(x).rstrip("0").rstrip(".") + "B", 10, color="#667085"))
    legend_y = height - 26
    methods = ["direct", "prompt", "qlora", "full"]
    for i, method in enumerate(methods):
        x = width / 2 - 210 + i * 140
        body.append(f'<line x1="{x}" y1="{legend_y-4}" x2="{x+24}" y2="{legend_y-4}" stroke="{COLORS[method]}" stroke-width="3"/>')
        body.append(svg_text(x + 32, legend_y, method, 12, anchor="start"))
    save_svg(path, width, height, body)


def draw_radar(path, axes, series, title):
    width, height = 720, 640
    cx, cy, r = width / 2, height / 2 + 12, 225
    body = [svg_text(width / 2, 36, title, 20, "700")]
    n = len(axes)
    for level in range(1, 6):
        pts = []
        rr = r * level / 5
        for i in range(n):
            ang = -math.pi / 2 + i * 2 * math.pi / n
            pts.append((cx + rr * math.cos(ang), cy + rr * math.sin(ang)))
        poly = " ".join(f"{x:.1f},{y:.1f}" for x, y in pts)
        body.append(f'<polygon points="{poly}" fill="none" stroke="{COLORS["grid"]}" stroke-width="1"/>')
    for i, axis in enumerate(axes):
        ang = -math.pi / 2 + i * 2 * math.pi / n
        x, y = cx + r * math.cos(ang), cy + r * math.sin(ang)
        body.append(f'<line x1="{cx}" y1="{cy}" x2="{x:.1f}" y2="{y:.1f}" stroke="{COLORS["grid"]}" stroke-width="1"/>')
        label_x, label_y = cx + (r + 44) * math.cos(ang), cy + (r + 44) * math.sin(ang)
        body.append(svg_text(label_x, label_y + 4, axis, 11))
    for name, vals, color in series:
        pts = []
        for i, v in enumerate(vals):
            ang = -math.pi / 2 + i * 2 * math.pi / n
            rr = r * max(0, min(v, 1))
            pts.append((cx + rr * math.cos(ang), cy + rr * math.sin(ang)))
        poly = " ".join(f"{x:.1f},{y:.1f}" for x, y in pts)
        body.append(f'<polygon points="{poly}" fill="{color}" fill-opacity="0.16" stroke="{color}" stroke-width="2.4"/>')
    lx = 76
    for i, (name, _, color) in enumerate(series):
        y = height - 38 + (i % 2) * 20
        x = lx + (i // 2) * 160
        body.append(f'<rect x="{x}" y="{y-12}" width="13" height="13" rx="3" fill="{color}"/>')
        body.append(svg_text(x + 20, y, name, 12, anchor="start"))
    save_svg(path, width, height, body)


def draw_slope(path, groups, title):
    width, height = 900, 560
    left_x, right_x = 245, 655
    top, step = 98, 54
    body = [svg_text(width / 2, 36, title, 20, "700")]
    body.append(svg_text(left_x, 72, "Before", 14, "700"))
    body.append(svg_text(right_x, 72, "After", 14, "700"))
    max_v = max(max(g["before"], g["after"]) for g in groups) or 1
    for i, g in enumerate(groups):
        y0 = top + i * step
        before = g["before"] / max_v
        after = g["after"] / max_v
        y_before = y0 - before * 18
        y_after = y0 - after * 18
        color = "#2E9F7D" if g["after"] >= g["before"] else "#C94545"
        body.append(svg_text(70, y0 + 5, g["label"], 12, anchor="start"))
        body.append(f'<line x1="{left_x}" y1="{y_before:.1f}" x2="{right_x}" y2="{y_after:.1f}" stroke="{color}" stroke-width="2.4"/>')
        body.append(f'<circle cx="{left_x}" cy="{y_before:.1f}" r="5" fill="{color}"/>')
        body.append(f'<circle cx="{right_x}" cy="{y_after:.1f}" r="5" fill="{color}"/>')
        body.append(svg_text(left_x - 12, y_before + 4, f"{g['before']:.3f}", 11, anchor="end"))
        body.append(svg_text(right_x + 12, y_after + 4, f"{g['after']:.3f}", 11, anchor="start"))
    save_svg(path, width, height, body)


def draw_scatter(path, rows, title):
    width, height = 880, 620
    left, right, top, bottom = 90, 50, 70, 78
    plot_w, plot_h = width - left - right, height - top - bottom
    body = [svg_text(width / 2, 36, title, 20, "700")]
    for i in range(6):
        x = left + plot_w * i / 5
        y = top + plot_h - plot_h * i / 5
        body.append(f'<line x1="{x:.1f}" y1="{top}" x2="{x:.1f}" y2="{top+plot_h}" stroke="{COLORS["grid"]}" stroke-width="1"/>')
        body.append(f'<line x1="{left}" y1="{y:.1f}" x2="{left+plot_w}" y2="{y:.1f}" stroke="{COLORS["grid"]}" stroke-width="1"/>')
        body.append(svg_text(x, top + plot_h + 26, f"{i/5:.1f}", 10, color="#667085"))
        body.append(svg_text(left - 14, y + 4, f"{i/5:.1f}", 10, anchor="end", color="#667085"))
    body.append(svg_text(left + plot_w / 2, height - 24, "Strict F1", 13, "700"))
    body.append(svg_text(24, top + plot_h / 2, "Contract Hit", 13, "700"))
    for r in rows:
        x = left + plot_w * max(0, min(r["x"], 1))
        y = top + plot_h - plot_h * max(0, min(r["y"], 1))
        color = COLORS.get(r["method"], "#333")
        body.append(f'<circle cx="{x:.1f}" cy="{y:.1f}" r="{r["size"]:.1f}" fill="{color}" fill-opacity="0.68" stroke="white" stroke-width="1.5"/>')
    for i, method in enumerate(["direct", "prompt", "qlora", "full"]):
        x = 555 + i * 78
        body.append(f'<circle cx="{x}" cy="{height-24}" r="6" fill="{COLORS[method]}"/>')
        body.append(svg_text(x + 11, height - 20, method, 11, anchor="start"))
    save_svg(path, width, height, body)


def draw_heatmap(path, families, methods, values, title):
    width, height = 920, 600
    left, top = 230, 78
    cell_w, cell_h = 145, 32
    body = [svg_text(width / 2, 36, title, 20, "700")]
    for j, method in enumerate(methods):
        body.append(svg_text(left + j * cell_w + cell_w / 2, top - 20, method, 12, "700"))
    for i, family in enumerate(families):
        body.append(svg_text(left - 14, top + i * cell_h + 21, family, 11, anchor="end"))
        for j, method in enumerate(methods):
            v = values.get((family, method))
            if v is None:
                color = "#EEF1F6"
                label = "-"
            else:
                t = max(0, min(v / 0.9, 1))
                r = int(244 - 154 * t)
                g = int(248 - 52 * t)
                b = int(251 - 118 * t)
                color = f"rgb({r},{g},{b})"
                label = f"{v:.2f}"
            x, y = left + j * cell_w, top + i * cell_h
            body.append(f'<rect x="{x}" y="{y}" width="{cell_w-4}" height="{cell_h-4}" rx="5" fill="{color}" stroke="white"/>')
            body.append(svg_text(x + cell_w / 2, y + 20, label, 10))
    save_svg(path, width, height, body)


def draw_framework(path):
    width, height = 1120, 440
    nodes = [
        ("Multi-source\nData", 70, 140),
        ("Governance\n& Splits", 245, 140),
        ("Tasks\nH/T/L", 420, 140),
        ("Modes\nD/P/FT", 595, 140),
        ("Models\n21 Open LLMs", 770, 140),
        ("Unified\nMetrics", 945, 140),
        ("Boundary\nAnalysis", 770, 285),
        ("Evidence\nSynthesis", 945, 285),
    ]
    body = [svg_text(width / 2, 36, "Unified Multi-task Evaluation Framework", 20, "700")]
    for i in range(5):
        x1 = nodes[i][1] + 116
        x2 = nodes[i + 1][1] - 10
        y = nodes[i][2] + 36
        body.append(f'<path d="M{x1},{y} L{x2},{y}" stroke="#9AA7BD" stroke-width="2" marker-end="url(#arrow)"/>')
    body.append(f'<path d="M1003,214 C1000,260 885,260 827,285" stroke="#9AA7BD" stroke-width="2" fill="none" marker-end="url(#arrow)"/>')
    body.append(f'<path d="M886,321 L934,321" stroke="#9AA7BD" stroke-width="2" marker-end="url(#arrow)"/>')
    defs = '<defs><marker id="arrow" markerWidth="10" markerHeight="10" refX="8" refY="3" orient="auto"><path d="M0,0 L0,6 L9,3 z" fill="#9AA7BD"/></marker></defs>'
    body.insert(0, defs)
    for label, x, y in nodes:
        body.append(f'<rect x="{x}" y="{y}" width="128" height="74" rx="13" fill="{COLORS["light"]}" stroke="#C9D2E3" stroke-width="1.4"/>')
        for k, line in enumerate(label.split("\n")):
            body.append(svg_text(x + 64, y + 30 + k * 18, line, 13, "700" if k == 0 else "400"))
    body.append(svg_text(width / 2, 405, "The framework emphasizes systematic evidence rather than a single-model leaderboard.", 14, "700", color="#667085"))
    save_svg(path, width, height, body)


def get_metric(row, task):
    if task == "has_vul":
        return fnum(row.get("f1"))
    if task == "vul_type":
        return fnum(row.get("macro_f1"))
    if task == "vul_line":
        return fnum(row.get("strict_f1"))
    return None


def top_rows(rows, task, metric, n=8):
    arr = [(fnum(r.get(metric)), r) for r in rows if r.get("任务") == task and fnum(r.get(metric)) is not None]
    arr.sort(key=lambda x: x[0], reverse=True)
    return arr[:n]


def method_means(rows):
    specs = [
        ("has_vul", "f1", "Has-F1"),
        ("vul_type", "macro_f1", "Type macro-F1"),
        ("vul_type", "multi_label_f1", "Type multi-label F1"),
        ("vul_line", "strict_f1", "Line strict-F1"),
        ("vul_line", "tolerant_f1", "Line tolerant-F1"),
        ("vul_line", "contract_hit", "Line contract-hit"),
    ]
    result = {}
    for task, metric, label in specs:
        by = defaultdict(list)
        for r in rows:
            if r.get("任务") == task and fnum(r.get(metric)) is not None:
                by[r.get("方法")].append(fnum(r.get(metric)))
        result[label] = {m: mean(v) for m, v in by.items()}
    return result


def main():
    FIG_DIR.mkdir(parents=True, exist_ok=True)
    wb = load_workbook(XLSX_PATH, data_only=True)
    detail = read_sheet(wb, "论文实验矩阵闭合明细")
    metrics = read_sheet(wb, "论文实验指标总表")
    summary = read_sheet(wb, "论文实验闭合汇总")
    prompt = read_sheet(wb, "Prompt消融")
    buqi = read_sheet(wb, "补齐前后12条明细")
    line_rows = read_sheet(wb, "line指标")

    completed = [r for r in detail if r.get("结果状态") == "有结果"]
    main_rows = [r for r in completed if r.get("系列") != "补齐前后对比"]
    model_count = len({r.get("模型") for r in main_rows if r.get("模型")})
    family_count = len({r.get("系列") for r in main_rows if r.get("系列")})
    formal_total = len(detail)
    formal_done = len(completed)
    formal_missing = formal_total - formal_done

    top_has = top_rows(main_rows, "has_vul", "f1", 6)
    top_type = top_rows(main_rows, "vul_type", "macro_f1", 6)
    top_line = top_rows(main_rows, "vul_line", "strict_f1", 6)
    means = method_means(main_rows)

    method_table_rows = []
    for method in ["direct", "prompt", "qlora", "full"]:
        method_table_rows.append([
            method,
            fmt(means["Has-F1"].get(method)),
            fmt(means["Type macro-F1"].get(method)),
            fmt(means["Type multi-label F1"].get(method)),
            fmt(means["Line strict-F1"].get(method)),
            fmt(means["Line tolerant-F1"].get(method)),
            fmt(means["Line contract-hit"].get(method)),
        ])

    best_table_rows = []
    for label, top, task, metric in [
        ("漏洞存在性", top_has, "has_vul", "f1"),
        ("漏洞类型", top_type, "vul_type", "macro_f1"),
        ("漏洞定位", top_line, "vul_line", "strict_f1"),
    ]:
        for rank, (v, r) in enumerate(top[:5], 1):
            best_table_rows.append([
                label if rank == 1 else "",
                rank,
                short_model(r.get("模型", "")),
                r.get("方法", ""),
                fmt(v),
                fmt(r.get("accuracy")),
                fmt(r.get("precision")),
                fmt(r.get("recall")),
                fmt(r.get("contract_hit")),
            ])

    # Figures.
    draw_framework(FIG_DIR / "fig1_framework.svg")
    labels = list(means.keys())
    bar_series = []
    for method in ["direct", "prompt", "qlora", "full"]:
        vals = [means[label].get(method, 0) for label in labels]
        bar_series.append((method, vals, COLORS[method]))
    draw_grouped_bar(FIG_DIR / "fig2_method_performance.svg", labels, bar_series, "Mode-level Mean Performance Across Tasks")

    axes = ["Has-F1", "Type-F1", "Type-macro", "Type-ML", "Line-strict", "Line-tolerant", "Hit"]
    radar_series = []
    for method in ["direct", "prompt", "qlora", "full"]:
        vals = [
            means["Has-F1"].get(method, 0),
            max([fnum(r.get("f1")) or 0 for r in main_rows if r.get("任务") == "vul_type" and r.get("方法") == method] or [0]),
            means["Type macro-F1"].get(method, 0),
            means["Type multi-label F1"].get(method, 0),
            means["Line strict-F1"].get(method, 0),
            means["Line tolerant-F1"].get(method, 0),
            means["Line contract-hit"].get(method, 0),
        ]
        radar_series.append((method, vals, COLORS[method]))
    draw_radar(FIG_DIR / "fig3_mode_radar.svg", axes, radar_series, "Capability Profile by Mode")

    qwen_rows = [r for r in main_rows if "Qwen2.5-Coder" in str(r.get("模型"))]
    x_sizes = sorted({param_size(r.get("模型", "")) for r in qwen_rows if param_size(r.get("模型", "")) is not None})
    panels = []
    for task, metric, title in [("has_vul", "f1", "Has-vul F1"), ("vul_type", "macro_f1", "Type macro-F1"), ("vul_line", "strict_f1", "Line strict-F1")]:
        panel = {"title": title, "x": x_sizes, "series": {}}
        for method in ["direct", "prompt", "qlora", "full"]:
            vals = []
            for x in x_sizes:
                rows = [r for r in qwen_rows if r.get("任务") == task and r.get("方法") == method and param_size(r.get("模型", "")) == x]
                vals.append((x, fnum(rows[0].get(metric)) if rows else None))
            panel["series"][method] = vals
        panels.append(panel)
    draw_line_panels(FIG_DIR / "fig4_qwen_scaling.svg", panels, "Qwen2.5-Coder Scaling Trends")

    # Prompt ablation in the manuscript focuses on the newer 7B/32B runs.
    # The 3B runs remain in the workbook as historical/background records.
    prompt_main_models = [
        ("7B", "Qwen2.5-Coder-7B-Instruct"),
        ("32B", "Qwen2.5-Coder-32B-Instruct"),
    ]
    p_labels = ["7B-Type", "32B-Type", "7B-Line", "32B-Line"]
    prompt_versions = ["P0", "P1", "P2", "P3"]
    p_series = []
    for pv in prompt_versions:
        vals = []
        for model_key, task, metric in [
            ("7B", "vul_type", "macro_f1"),
            ("32B", "vul_type", "macro_f1"),
            ("7B", "vul_line", "strict_f1"),
            ("32B", "vul_line", "strict_f1"),
        ]:
            candidates = [r for r in prompt if model_key in str(r.get("模型")) and r.get("任务") == task and r.get("版本") == pv]
            vals.append(fnum(candidates[0].get(metric)) if candidates else 0)
        p_series.append((pv, vals, ["#667085", "#D98C40", "#2E9F7D", "#C94545"][prompt_versions.index(pv)]))
    draw_grouped_bar(FIG_DIR / "fig5_prompt_ablation.svg", p_labels, p_series, "Structured Prompt Ablation")

    groups = []
    pair_keys = [
        ("Qwen-3B type", "Qwen/Qwen2.5-Coder-3B-Instruct", "vul_type", "qlora", "f1"),
        ("Qwen-3B line", "Qwen/Qwen2.5-Coder-3B-Instruct", "vul_line", "qlora", "strict_f1"),
        ("CodeBERT type", "microsoft/codebert-base", "vul_type", "full", "f1"),
        ("CodeBERT line", "microsoft/codebert-base", "vul_line", "full", "strict_f1"),
        ("GraphCodeBERT type", "microsoft/graphcodebert-base", "vul_type", "full", "f1"),
        ("GraphCodeBERT line", "microsoft/graphcodebert-base", "vul_line", "full", "strict_f1"),
    ]
    for label, model, task, method, metric in pair_keys:
        before = [r for r in buqi if r.get("模型") == model and r.get("任务") == task and r.get("方法") == method and r.get("补齐阶段说明") == "补齐前"]
        after = [r for r in buqi if r.get("模型") == model and r.get("任务") == task and r.get("方法") == method and r.get("补齐阶段说明") == "补齐后"]
        if before and after:
            groups.append({"label": label, "before": fnum(before[0].get(metric)) or 0, "after": fnum(after[0].get(metric)) or 0})
    draw_slope(FIG_DIR / "fig6_data_completion_slope.svg", groups, "Before vs After Data Completion")

    scatter_rows = []
    for r in main_rows:
        if r.get("任务") == "vul_line" and fnum(r.get("strict_f1")) is not None and fnum(r.get("contract_hit")) is not None:
            ps = param_size(r.get("模型", "")) or 1.0
            scatter_rows.append({"x": fnum(r.get("strict_f1")), "y": fnum(r.get("contract_hit")), "method": r.get("方法"), "size": 4 + min(math.log(ps + 1) * 2.2, 8)})
    draw_scatter(FIG_DIR / "fig7_line_boundary_scatter.svg", scatter_rows, "Line Localization: Exactness vs Contract-level Hit")

    families = sorted({r.get("系列") for r in main_rows if r.get("系列")})
    methods = ["direct", "prompt", "qlora", "full"]
    heat_values = {}
    for fam in families:
        for method in methods:
            vals = []
            for r in main_rows:
                if r.get("系列") == fam and r.get("方法") == method:
                    task = r.get("任务")
                    metric = {"has_vul": "f1", "vul_type": "macro_f1", "vul_line": "tolerant_f1"}.get(task)
                    if metric and fnum(r.get(metric)) is not None:
                        vals.append(fnum(r.get(metric)))
            heat_values[(fam, method)] = mean(vals) if vals else None
    draw_heatmap(FIG_DIR / "fig8_family_mode_heatmap.svg", families, methods, heat_values, "Family-mode Mean Performance Heatmap")

    literature = [
        ("GPT-4o real-world sample", 0.019),
        ("Detect Llama", 0.680),
        ("GPT-3.5FT", 0.776),
        ("Our best has_vul", fnum(top_has[0][0])),
        ("SmartConDetect", 0.909),
        ("SmartBugBERT", 0.912),
        ("G-Scan line", 0.937),
        ("SCVDIE", 0.976),
    ]
    draw_grouped_bar(
        FIG_DIR / "fig9_literature_anchor.svg",
        [x[0] for x in literature],
        [("F1", [x[1] for x in literature], "#3B82F6")],
        "Non-equivalent Literature Anchors",
        "Reported F1",
    )

    # Tables.
    closure_rows = []
    for r in summary:
        if str(r.get("模型")) == "总计":
            continue
        if fnum(r.get("缺结果数")) and fnum(r.get("缺结果数")) > 0:
            closure_rows.append([r.get("模型"), int(fnum(r.get("应跑任务数"))), int(fnum(r.get("已有结果数"))), int(fnum(r.get("缺结果数"))), r.get("缺失任务名称")])
    closure_table = md_table(["模型", "应跑", "已有", "缺口", "缺失任务"], closure_rows)

    task_headers = ["任务", "样本/测试规模", "主指标", "辅助指标", "论文解释重点"]
    task_rows = [
            ["`has_vul`", "test=2363", "F1 / Accuracy", "Precision, Recall", "二分类成熟度与微调收益"],
            ["`vul_type`", "test=1992", "macro-F1 / multi-label F1", "Accuracy, weighted-F1", "长尾类别与多标签能力边界"],
            ["`vul_line`", "test=807", "strict-F1 / tolerant-F1", "contract-hit, candidate-recall", "精确行定位与候选区域命中差异"],
    ]
    task_table = md_table(task_headers, task_rows)
    method_headers = ["方法", "Has-F1", "Type macro-F1", "Type multi-label F1", "Line strict-F1", "Line tolerant-F1", "Line hit"]
    method_table = md_table(method_headers, method_table_rows)
    best_headers = ["任务", "排名", "模型", "方法", "主指标", "Accuracy", "Precision", "Recall", "Contract-hit"]
    best_table = md_table(best_headers, best_table_rows)

    type_rows = []
    for rank, (v, r) in enumerate(top_type[:8], 1):
        type_rows.append([rank, short_model(r.get("模型")), r.get("方法"), fmt(r.get("f1")), fmt(r.get("macro_f1")), fmt(r.get("multi_label_f1")), fmt(r.get("accuracy"))])
    type_headers = ["排名", "模型", "方法", "F1", "Macro-F1", "Multi-label F1", "Accuracy"]
    type_table = md_table(type_headers, type_rows)

    line_top_rows = []
    for rank, (v, r) in enumerate(top_line[:8], 1):
        line_top_rows.append([rank, short_model(r.get("模型")), r.get("方法"), fmt(r.get("strict_f1")), fmt(r.get("tolerant_f1")), fmt(r.get("contract_hit")), fmt(r.get("candidate_recall"))])
    line_headers = ["排名", "模型", "方法", "Strict-F1", "Tolerant-F1", "Contract-hit", "Candidate-recall"]
    line_table = md_table(line_headers, line_top_rows)

    prompt_rows = []
    prompt_main = [
        r for r in prompt
        if any(model_name in str(r.get("模型")) for _, model_name in prompt_main_models)
    ]
    for r in prompt_main:
        metric = "macro_f1" if r.get("任务") == "vul_type" else "strict_f1"
        aux = "multi_label_f1" if r.get("任务") == "vul_type" else "contract_hit"
        prompt_rows.append([short_model(str(r.get("模型"))), r.get("任务"), r.get("版本"), r.get("模式"), fmt(r.get(metric)), fmt(r.get(aux)), fmt(r.get("parseable_ratio")), fmt(r.get("empty_output_rate"))])
    prompt_headers = ["模型", "任务", "Prompt", "模式", "主指标", "辅助指标", "Parseable", "Empty"]
    prompt_table = md_table(prompt_headers, prompt_rows)

    buqi_rows = []
    for label, model, task, method, metric in pair_keys:
        before = [r for r in buqi if r.get("模型") == model and r.get("任务") == task and r.get("方法") == method and r.get("补齐阶段说明") == "补齐前"]
        after = [r for r in buqi if r.get("模型") == model and r.get("任务") == task and r.get("方法") == method and r.get("补齐阶段说明") == "补齐后"]
        if before and after:
            before_v, after_v = fnum(before[0].get(metric)) or 0, fnum(after[0].get(metric)) or 0
            buqi_rows.append([label, metric, fmt(before_v), fmt(after_v), fmt(after_v - before_v), before[0].get("test_examples"), after[0].get("test_examples")])
    buqi_headers = ["配对实验", "指标", "补齐前", "补齐后", "变化", "前测样本", "后测样本"]
    buqi_table = md_table(buqi_headers, buqi_rows)

    literature_headers = ["文献/系统", "任务口径", "报告值", "与本文可比性说明"]
    literature_rows = [
            ["Detect Llama", "LLM 二分类", "F1=0.680", "与本文 `has_vul` 同属 LLM 检测方向，但数据集不同"],
            ["GPT-3.5FT", "LLM 二分类", "F1=0.776", "可作为微调 LLM 的弱横向锚点"],
            ["SmartConDetect", "BERT/静态分析检测", "F1=0.909", "专门检测框架，数据集与切片策略不同"],
            ["SmartBugBERT", "Bytecode+CFG 检测", "F1=0.9119", "工具标注数据，非本文多任务矩阵"],
            ["G-Scan", "重入漏洞行级定位", "F1=0.9369", "单漏洞、高质量图结构定位，不宜与本文多类多源定位直接等同"],
            ["SCVDIE", "信息图+集成学习", "F1=0.9757", "专门集成模型，通常是单一任务或特定数据协议"],
            ["本文最佳 `has_vul`", "开源 LLM 全参微调", "F1=0.8757", "低于部分专门高性能检测器，但高于部分 LLM 检测基线"],
            ["本文最佳 `vul_line`", "多类多源行级定位", "strict-F1=0.2955 / hit=0.8451", "用于刻画精确行定位与候选区域命中的差异"],
    ]
    literature_table = md_table(literature_headers, literature_rows)

    # Compose draft.
    fig = lambda name: f"figures_q1/{name}"
    highlights = """# Highlights

- Open source language models are evaluated across three smart contract security tasks
- Fine tuning consistently outperforms direct and prompt based use modes
- Vulnerability localization shows high contract hits but limited strict line accuracy
- Dataset completion changes task difficulty and exposes long tail weaknesses
- The benchmark clarifies model capability boundaries for software security practice
"""

    draft = f"""# Systematic Evaluation of Open Source Large Language Models for Multitask Smart Contract Vulnerability Analysis

## Structured Abstract

**Context:** Smart contracts execute business logic in immutable blockchain environments, where software defects can cause severe and irreversible losses. Open source large language models have shown promise for code understanding and security analysis, but existing studies rarely evaluate vulnerability detection, vulnerability type identification, and vulnerable line localization under a unified software engineering protocol.

**Objectives:** This study characterizes the capability boundaries of open source large language models for smart contract vulnerability analysis and explains how task type, training mode, model family, prompt design, and task oriented dataset completion affect empirical conclusions.

**Methods:** We construct a multitask benchmark with unified data splits, output protocols, and evaluation metrics. The experiment compares {model_count} models from {family_count} model families across direct inference, structured prompting, full fine tuning, and QLoRA. The analysis covers binary detection, type classification, line localization, prompt ablation, model scaling, and dataset completion comparisons for type and line tasks.

**Results:** The results show a clear task hierarchy. Binary vulnerability detection is the most mature task, with the best F1 reaching {fmt(top_has[0][0])}. Vulnerability type classification obtains a best ordinary F1 of {fmt(top_rows(main_rows, 'vul_type', 'f1', 1)[0][0])}, but the best macro F1 is only {fmt(top_type[0][0])} and the best multi label F1 is {fmt(top_rows(main_rows, 'vul_type', 'multi_label_f1', 1)[0][0])}. Vulnerable line localization remains the most difficult task, with a best strict F1 of {fmt(top_line[0][0])} but a contract hit rate of {fmt(top_rows(main_rows, 'vul_line', 'contract_hit', 1)[0][0])}. Fine tuning is the most stable source of improvement, while prompt and scale effects are not monotonic.

**Conclusion:** Open source large language models are promising trainable components for smart contract vulnerability analysis, but they should not be treated as universal replacements for specialized analyzers. Their practical value lies in configurable and reproducible use within data governed software security workflows, especially when line localization is supported by structural program information and post processing mechanisms.

**Keywords:** Smart contracts; Vulnerability detection; Large language models; Empirical software engineering; Software security; Vulnerability localization; Fine tuning

## Highlights

- Open source language models are evaluated across three smart contract security tasks
- Fine tuning consistently outperforms direct and prompt based use modes
- Vulnerability localization shows high contract hits but limited strict line accuracy
- Dataset completion changes type and line task difficulty and exposes long tail weaknesses
- The benchmark clarifies model capability boundaries for software security practice

## 1 引言

智能合约以可编程方式执行链上资产转移与业务逻辑，其不可篡改性和公开可调用性使漏洞后果具有高风险和高扩散特征。传统智能合约漏洞检测主要依赖静态分析、符号执行、模糊测试和形式化验证等方法，这些方法具备较好的可解释性，但在跨函数依赖、业务逻辑漏洞、低频类别和复杂交互场景中仍面临规则覆盖不足、路径爆炸和误报偏高等问题。随着代码预训练模型和开源大语言模型的发展，研究者开始尝试使用 CodeBERT、GraphCodeBERT、CodeT5、Code Llama、Qwen2.5-Coder、DeepSeek-Coder 等模型进行漏洞检测、漏洞分类和漏洞解释。然而，仅报告单个模型在单个任务上的最高分，已经难以回答智能合约安全实践真正关心的问题：开源大模型究竟在哪类任务上可靠，在哪类任务上只是看似有用；prompt 工程是否真的稳定有效；微调收益是否足以抵消训练成本；模型规模增加是否必然带来更好性能；多源数据补齐是否一定提升所有任务。

本文从系统评测视角出发，研究重点不是提出新的检测网络结构，而是建立面向开源大语言模型的智能合约多任务漏洞分析基准。该视角与当前实验现象相一致：本文的二分类检测结果已经具备较强竞争力，但漏洞类型宏平均和行级精确定位仍显著落后于部分专门图模型或规则增强方法。因此，本文的核心价值在于建立可复现的实验矩阵、统一的任务协议和细粒度分析框架，以揭示开源大模型在不同任务层级上的真实能力边界。

本文主要回答四个研究问题：

- `RQ1`：开源大语言模型在漏洞存在性判断、漏洞类型识别和漏洞行定位三类任务上的能力层级如何？
- `RQ2`：direct、prompt、full fine-tuning 和 QLoRA 四类使用模式的性能差异和稳定性如何？
- `RQ3`：不同模型系列和参数规模对三类任务的影响是否一致？
- `RQ4`：多源数据补齐、结构化 prompt 和定位指标设计如何改变实验结论的解释方式？

本文的主要贡献包括四点。第一，构建覆盖三类漏洞分析任务的统一评测框架，并将开源 LLM、传统代码预训练模型和不同训练模式纳入同一实验矩阵。第二，建立面向多源异构智能合约数据的数据治理协议，显式处理标签映射、低频类别补齐、去重切分和行级监督质量控制。第三，重新定义漏洞行定位任务为多行预测问题，报告 strict-F1、tolerant-F1、contract-hit 和 candidate-recall 等互补指标，避免单一指标误导。第四，基于闭合实验矩阵进行多角度分析，揭示 prompt 工程不稳定、监督微调主导性能提升、模型规模收益非单调和行级定位仍存在明显边界等结论。

![统一多任务评测框架]({fig('fig1_framework.svg')})

**图1 统一多任务评测框架。** 本文将多源数据治理、三类任务、四种模式和统一指标连接为一条可复现评测链路，研究目标从单点性能比较扩展为系统化能力边界分析。

## 2 相关工作

### 2.1 智能合约漏洞检测

早期智能合约漏洞检测主要依赖静态分析和符号执行。Oyente、Securify、Slither 等工具通过路径约束、语义规则和中间表示检测重入、时间依赖、未检查调用和访问控制等问题。此类工具的优势在于可解释性强、工程集成成本较低，但其检测能力通常受限于预定义规则和路径搜索策略。后续研究引入图神经网络、序列模型和专家模式融合方法，例如 CGE、MANDO 和 G-Scan 等工作，通过控制流、数据流、调用图和异构图结构捕获更复杂的语义依赖。

这些方法往往能够在特定漏洞类型或特定数据集上取得较高数值，但它们也存在两个局限。其一，不同研究使用的数据来源、标签定义、切分策略和评价指标差异较大，横向比较困难。其二，大量工作聚焦于单一任务，例如合同级二分类或重入漏洞定位，而缺少覆盖存在性判断、类型识别和行级定位的统一多任务分析。本文正是在这一背景下，将“可比评测”和“能力边界解释”作为核心目标。

### 2.2 代码预训练模型与开源大语言模型

CodeBERT、GraphCodeBERT、CodeT5 和 UniXcoder 等代码预训练模型为漏洞检测提供了通用语义表示。与传统静态工具相比，预训练模型能够从数据中学习语义模式，适合构建监督分类器。近年来，Code Llama、Qwen2.5-Coder、DeepSeek-Coder、StarCoder2、CodeGemma、Granite Code 和 WizardCoder 等开源大模型进一步扩展了代码理解和生成能力，使研究者能够在本地环境中开展 direct 推理、结构化 prompt、全参微调和参数高效微调。

然而，大语言模型并不天然等价于高精度安全检测器。已有 LLM 漏洞检测工作表明，prompt 输出可能受到格式解析、幻觉、类别不平衡和上下文截断的影响；真实复杂合约上的结果也可能显著低于基准样本。本文因此不默认 prompt 有效，而是将 prompt 作为一个需要被单独消融和诊断的实验变量。

### 2.3 研究空缺

现有研究至少留下三个空缺。第一，缺少同一协议下的多任务对比，尤其缺少将漏洞存在性、漏洞类型和漏洞定位放在同一矩阵中的研究。第二，缺少对 direct、prompt、full 和 QLoRA 等模式的系统比较，导致模型能力、提示工程收益和监督微调收益难以拆分。第三，行级定位任务的指标口径不统一，部分研究只报告命中或排名指标，难以同时解释精确行定位和候选区域命中之间的差异。本文围绕这些空缺展开。

## 3 任务定义与评测框架

本文定义三类任务。`has_vul` 是漏洞存在性二分类任务，输入为合约或函数级代码片段，输出是否存在漏洞。`vul_type` 是漏洞类型识别任务，目标标签空间包括 `access_control`、`arithmetic`、`bad_randomness`、`denial_service`、`front_running`、`reentrancy`、`time_manipulation` 和 `unchecked_low_calls` 等类别。由于真实智能合约可能包含多个漏洞，本文保留多标签辅助评估，而不是强行将样本压缩为单标签。`vul_line` 是漏洞行定位任务，输出为一个或多个候选漏洞行，评价时同时考虑精确行命中、邻域容忍命中和合约级候选命中。

**表1 三类任务定义与评价指标。**

{task_table}

四类模式定义如下。`direct` 表示不训练、不加入复杂任务模板的直接推理；`prompt` 表示加入显式任务说明、输出格式约束和结构化模板；`full` 表示全参数监督微调；`qlora` 表示参数高效微调。本文将 prompt 与 fine-tune 明确区分，因为二者体现了完全不同的工程代价和能力来源。

## 4 数据集构建与质量控制

本文以公开智能合约漏洞数据和项目实验数据为基础，构建统一数据治理流程。数据来源包括 Empirical-Analysis-of-Vulnerability-Detection-Tools-for-Solidity-Smart-Contracts、SolidiFI、DAppSCAN、Slither Audited Smart Contracts、ScrawlD 和 smart-contract-vulndb 等。不同来源在粒度、标签可靠性、源代码可得性和行级监督质量上差异较大，因此本文采用“统一标签空间 + 任务特定接入协议”的方式组织数据，而不是简单拼接。

在数据切分上，本文采用训练、验证和测试划分，并结合来源分层、类别分层和代码级去重，降低训练测试泄漏风险。对于 `vul_type`，无法稳定映射到目标标签空间的样本被过滤；对于 `vul_line`，越界行号和弱定位标签被清理。对于低频类别，本文通过候选生成、人工/规则审核和回流补齐方式扩展 `bad_randomness`、`denial_service` 和 `front_running` 等类别。

**表2 数据集规模与任务角色。**

| 任务 | Train | Val | Test | 任务角色 |
| --- | ---: | ---: | ---: | --- |
| `has_vul` | 17411 | 4667 | 2363 | 二分类存在性判断 |
| `vul_type` | 24992 | 3829 | 1992 | 8 类漏洞类型识别与多标签辅助评估 |
| `vul_line` | 10155 | 1529 | 807 | 多行漏洞定位与候选区域命中评估 |

## 5 实验设计

本文主分析仅纳入具有完整评价指标的正式实验记录。正式结果来自 `论文实验矩阵闭合明细` 和 `论文实验指标总表`，补齐前后实验单独来自 `补齐前后12条明细`。这一统计口径用于避免将重复 run、历史调参 run 或不完整记录混入主分析，从而保证模型、任务和方法之间的比较具有一致性。

**表3 正式实验闭合缺口。**

{closure_table}

模型覆盖 Qwen2.5-Coder、DeepSeek-Coder、StarCoder2、CodeGemma、Granite Code、CodeLlama、WizardCoder 等开源代码大模型，同时包含 CodeBERT、GraphCodeBERT、UniXcoder 和 CodeT5 等传统代码预训练基准。对于 0.5B 至 3B 模型，尽量覆盖 direct、prompt、full 和 QLoRA；对于 7B 及以上模型，受显存和训练成本限制，主要覆盖 direct、prompt 和 QLoRA。评价指标按任务区分：`has_vul` 使用 accuracy、precision、recall 和 F1；`vul_type` 额外报告 macro-F1 和 multi-label F1；`vul_line` 报告 strict-F1、tolerant-F1、contract-hit 和 candidate-recall。

## 6 实验结果

### 6.1 总体结果与任务难度层级

从最佳结果看，三类任务呈现明显难度层级。`has_vul` 的最佳 F1 为 {fmt(top_has[0][0])}，说明在统一数据协议和监督微调条件下，开源代码大模型已经能够较稳定完成漏洞存在性判断。`vul_type` 的最佳普通 F1 为 {fmt(top_rows(main_rows, 'vul_type', 'f1', 1)[0][0])}，但最佳 macro-F1 仅为 {fmt(top_type[0][0])}，说明总体分类能力与长尾类别均衡识别之间存在明显差距。`vul_line` 的最佳 strict-F1 为 {fmt(top_line[0][0])}，tolerant-F1 为 {fmt(top_rows(main_rows, 'vul_line', 'tolerant_f1', 1)[0][0])}，contract-hit 为 {fmt(top_rows(main_rows, 'vul_line', 'contract_hit', 1)[0][0])}，表明模型具有较强候选区域检索能力，但精确行级定位仍是瓶颈。

**表4 三类任务最佳结果概览。**

{best_table}

![不同模式的任务平均表现]({fig('fig2_method_performance.svg')})

**图2 不同模式在多任务上的平均表现。** full fine-tuning 在存在性判断和类型识别上明显领先；QLoRA 通常优于 direct 和 prompt，但在行级严格定位上仍存在明显上限。

### 6.2 不同模式的收益差异

模式对比是本文最稳定的结论之一。direct 和 prompt 在三类任务上整体较弱，尤其在 `vul_type` 和 `vul_line` 上难以形成可靠输出。full fine-tuning 在 `has_vul` 和 `vul_type` 上优势明显，QLoRA 则在资源受限条件下提供了较好的折中。以均值来看，`has_vul` 的 direct 和 prompt F1 仅为 0.1599 和 0.1885，而 QLoRA 和 full 分别达到 0.7322 和 0.8564。`vul_type` 的 macro-F1 从 direct/prompt 的 0.0863/0.0532 提升到 QLoRA/full 的 0.3849/0.4407。`vul_line` 的 tolerant-F1 也从 direct/prompt 的 0.1053/0.1132 提升到 QLoRA/full 的 0.2960/0.3886。

**表5 不同运行模式的平均性能。**

{method_table}

![模式能力雷达图]({fig('fig3_mode_radar.svg')})

**图3 模式能力画像。** direct 与 prompt 在多任务维度上均较弱，监督微调显著扩大了能力边界；full 在分类类任务上最稳，QLoRA 在低资源条件下具有较好的可行性。

### 6.3 漏洞存在性判断

漏洞存在性判断是当前结果中最成熟的任务。最佳模型为 Qwen2.5-Coder-1.5B full，F1 为 {fmt(top_has[0][0])}；Qwen2.5-Coder-0.5B full、Granite-3B full 和 Qwen2.5-Coder-3B full 的 F1 也非常接近。这说明对于二分类检测任务，较小规模的代码模型在充分监督下已经可以达到较强水平，参数规模并不是唯一决定因素。

这一结果表明，开源代码大模型在二分类场景中已经具备较好的监督适配能力。与部分 LLM 检测基线相比，本文 `has_vul` 结果具有竞争力；但与专门图结构或规则增强模型相比，仍未达到 0.90 以上的高分区间。因此，本文将二分类结果解释为开源大模型可训练性和可部署性的证据，而非单点最优模型声明。

### 6.4 漏洞类型识别

漏洞类型识别的结果显示，模型可以学习主流类别，但对长尾类别和多标签样本仍不稳定。表5给出了 macro-F1 排名前列的结果。可以看到，小模型 full 微调表现非常强，Qwen-0.5B full 和 Granite-3B full 均进入前列；但 macro-F1 最高仅为 {fmt(top_type[0][0])}，明显低于普通 F1。这说明多数类别上的平均能力仍受到类别不平衡和低频漏洞语义复杂性的限制。

**表6 漏洞类型识别 Top 结果。**

{type_table}

这一结果恰恰构成本文的一个重要发现：如果只报告普通 F1 或 accuracy，会高估模型对漏洞类型空间的整体掌握程度；macro-F1 和 multi-label F1 才能揭示模型对低频类别和多漏洞样本的真实能力。因此，本文在解释 `vul_type` 结果时同时强调总体分类性能、类别均衡性能和多标签识别能力。

### 6.5 漏洞行定位

漏洞行定位是本文最能体现能力边界的任务。表6显示，最佳 strict-F1 约为 {fmt(top_line[0][0])}，明显低于二分类和类型识别；但 contract-hit 可以达到 {fmt(top_rows(main_rows, 'vul_line', 'contract_hit', 1)[0][0])}。这意味着模型并非完全不能理解漏洞位置，而是常常能够把候选范围缩小到相关合约或代码区域，却难以稳定给出完全精确的行号。

**表7 漏洞行定位 Top 结果。**

{line_table}

![行级定位边界散点图]({fig('fig7_line_boundary_scatter.svg')})

**图4 行级定位的精确性与合约级命中。** 横轴为 strict-F1，纵轴为 contract-hit。大量点呈现“hit 高、strict-F1 低”的形态，说明模型更擅长定位相关区域，而非精确行号。

这一结论表明，`vul_line` 的性能瓶颈并非完全来自模型无法理解漏洞位置，而是来自候选区域识别与精确行号判定之间的差距。开源大模型在行级定位中已经具备一定候选区域识别能力，但从候选区域到精确漏洞行仍需要结构化程序信息、排序损失、行级监督和后处理机制进一步增强。

## 7 深入分析

### 7.1 模型规模并非单调收益

Qwen2.5-Coder 系列提供了观察参数规模影响的良好切面。图5展示了从 0.5B 到 32B 的趋势。总体看，direct 模式下较大模型通常更好，但在监督微调后，小模型并不必然落后。尤其在 `has_vul` 和 `vul_type` 上，0.5B、1.5B 和 3B full 微调已经达到非常接近的高分，说明领域监督数据和任务适配比单纯参数规模更关键。

![Qwen 系列规模趋势]({fig('fig4_qwen_scaling.svg')})

**图5 Qwen2.5-Coder 系列规模趋势。** 规模收益并不单调，尤其在监督微调条件下，小模型可达到接近甚至超过大模型 QLoRA 的效果。

这一发现对工程部署有实际意义：如果目标是低成本二分类或类型识别，小模型 full fine-tuning 可能比大模型 prompt 或大模型 QLoRA 更具性价比；如果目标是泛化推理或更复杂上下文理解，大模型仍可能提供更高上限，但需要额外实验验证。

### 7.2 结构化 Prompt 消融

Prompt 消融结果显示，结构化 prompt 的收益并不稳定。不同模型、不同任务和不同模板之间并不存在统一最优模板。本文消融采用 Qwen2.5-Coder-7B 与 Qwen2.5-Coder-32B 两组结果，共覆盖 2 个模型、2 个任务和 4 种 prompt 模板，合计 16 组实验。在部分 `vul_type` 实验中，P0 或 P1 反而优于更复杂的结构化模板；在 `vul_line` 中，结构化输出要求可能增加解析失败或空输出比例。表8给出了 7B/32B 的 Prompt 消融结果。

**表8 结构化 Prompt 消融结果。**

{prompt_table}

![Prompt 消融图]({fig('fig5_prompt_ablation.svg')})

**图6 结构化 Prompt 消融。** Prompt 模板收益依赖模型和任务，复杂模板并不必然提升结果。对行级定位而言，格式约束还可能带来空输出或解析失败。

因此，本文应避免将 prompt 工程描述为稳定增益来源。更准确的结论是：prompt 是一种低成本适配手段，但其效果受到输出格式、解析器、类别空间和模型指令遵循能力的共同影响；对于严肃漏洞检测，prompt 不能替代监督微调。

### 7.3 数据补齐前后对比

数据补齐实验进一步说明，数据规模扩大并不必然在所有任务上带来同方向收益。表9展示了 6 组补齐前后配对结果。该分析聚焦于 `vul_type` 和 `vul_line`，原因是本轮数据补齐主要改变漏洞类型覆盖、低频类别分布和行级定位监督质量；`has_vul` 是二分类存在性判断任务，其标签空间不会随低频类型补齐和行号补齐发生同质变化，因此放入同一组补齐前后配对分析容易把任务粒度差异误读为数据补齐收益。对于 `vul_type`，补齐后测试样本规模扩大，accuracy 往往提升，但 F1 或 multi-label 指标可能因类别分布变化而出现波动；对于 `vul_line`，补齐后 strict-F1 下降，但这是在更大、更复杂测试集和更严格定位监督下发生的，不能简单解释为模型退化。

**表9 数据补齐前后配对结果。**

{buqi_table}

![补齐前后斜率图]({fig('fig6_data_completion_slope.svg')})

**图7 数据补齐前后对比。** 数据补齐改变了任务难度和测试分布，分类与定位任务受到的影响方向并不完全一致。

该结果说明，数据补齐的作用不能被简化为单向性能提升。数据扩展既能改善类别覆盖，也可能通过引入更复杂样本暴露模型在低频类别和精确定位上的不足，因此更适合作为数据治理效应来解释。

### 7.4 模型系列与方法热力图

图8从模型系列和方法两个维度展示平均表现。可以看到，不同系列之间的差异并不只由参数规模决定，预训练语料、代码建模目标、指令遵循能力和微调稳定性都会影响结果。同一系列内部，full 或 QLoRA 通常明显优于 direct 和 prompt；不同系列之间，在相近参数规模下也存在显著波动。

![模型系列-方法热力图]({fig('fig8_family_mode_heatmap.svg')})

**图8 模型系列与模式的平均表现热力图。** 颜色越深表示平均表现越高。该图强调模型家族、训练模式和任务协议共同决定最终效果。

### 7.5 与代表性文献的数值锚点对比

需要特别说明，不同智能合约漏洞检测论文的数据集、标签口径、漏洞类型、任务粒度和评价指标差异很大，因此不能把分数直接等价比较。表9仅提供数值锚点，用于判断本文结果处于什么区间。

**表10 代表性文献数值锚点。**

{literature_table}

![代表性文献数值锚点]({fig('fig9_literature_anchor.svg')})

**图9 非等价文献数值锚点。** 图中数值来自不同数据集和任务，用于帮助判断本文结果的大致区间，而非进行直接胜负比较。

这些数值锚点表明，本文结果处于一个较为清晰的位置：`has_vul` 已经高于部分 LLM 检测基线；`vul_type` 的 macro-F1 和 multi-label F1 反映了多类长尾任务的真实难度；`vul_line` 的 strict-F1 虽低，但 contract-hit 较高，支持“模型能够定位相关区域，但精确行级判断仍受限”的能力边界结论。由此可见，本文的主要价值并非替代已有专门检测器，而是为开源大模型在智能合约漏洞分析中的适用范围和失败模式提供系统证据。

## 8 讨论

### 8.1 为什么二分类强而行级定位弱

二分类任务只要求模型判断是否存在漏洞，其监督信号相对粗粒度，模型可以利用语义、模式和统计线索形成有效判断。行级定位则要求模型在大量非漏洞行中识别少数关键行，且真实漏洞可能跨越多行、多函数甚至依赖外部调用关系。因此，行级 strict-F1 低并不意外。更有解释力的是 contract-hit 和 tolerant-F1，它们说明模型是否掌握了大致位置。本文结果显示，开源大模型已经具备一定候选区域识别能力，但还缺少精确定位所需的程序结构归纳能力。

### 8.2 为什么小模型 full 能超过大模型 prompt

大模型 direct/prompt 依赖预训练知识和指令遵循能力，而漏洞检测是一个强领域、强标签、强格式约束任务。没有监督适配时，大模型容易输出不稳定解释或泛化安全常识，无法稳定映射到目标标签空间。小模型 full fine-tuning 虽然参数少，但直接学习了任务标签边界，因此在分类任务中表现突出。这一发现提示，智能合约漏洞检测并不是单纯堆叠参数规模的问题。

### 8.3 对基准研究的启示

本文结果表明，面向安全任务的开源大模型评测应从单一得分比较转向多维证据综合。首先，二分类、类型识别和行级定位之间存在明显任务层级差异，单个任务的高分不能代表模型具备完整审计能力。其次，prompt、QLoRA 和全参微调反映了不同的能力来源和工程代价，应在同一数据协议下共同分析。最后，数据补齐和标签治理会改变任务难度与评价分布，因此基准研究需要同时报告数据来源、切分策略、标签映射和错误类型，而不只是最终指标。

## 9 有效性威胁

本文仍存在若干有效性威胁。第一，尽管进行了代码级去重和统一切分，多源数据仍可能存在来源偏差和标签噪声。第二，部分漏洞类型的长尾分布仍然明显，macro-F1 对少数类波动较敏感。第三，行级定位标签来自多源数据和规则补齐，虽然经过清洗，但仍难完全等价于人工审计中的漏洞根因定位。第四，不同文献之间的数据集、任务粒度和评价指标差异较大，本文的外部数值对比仅作为区间锚点，不作为直接胜负结论。

## 10 结论

本文围绕开源大语言模型在智能合约漏洞分析中的真实能力边界，构建了覆盖漏洞存在性判断、漏洞类型识别和漏洞行定位三类任务的统一评测框架，并系统比较 direct、prompt、full fine-tuning 和 QLoRA 等模式。实验表明，开源大模型在二分类检测上已经具备较强可用性，在类型识别上受到长尾类别和多标签分布限制，在行级定位上则表现出“候选区域可命中、精确行号仍困难”的典型边界。监督微调是最稳定的性能来源，结构化 prompt 的收益不稳定，参数规模收益也并非单调。

这些发现说明，当前阶段不宜将开源大语言模型简单描述为智能合约漏洞检测的通用替代方案。更合理的结论是：开源大模型适合作为可训练、可部署、可组合的漏洞分析组件，但仍需要数据治理、结构化程序表示、行级监督和后处理机制共同支撑。未来工作将进一步探索将图结构、静态分析候选和大模型微调结合的混合式漏洞定位框架。

## References

[1] Luu, L., Chu, D. H., Olickel, H., Saxena, P., and Hobor, A. Making Smart Contracts Smarter. In *Proceedings of ACM CCS*, 2016.

[2] Tsankov, P., Dan, A., Drachsler-Cohen, D., Gervais, A., Buenzli, F., and Vechev, M. Securify: Practical Security Analysis of Smart Contracts. In *Proceedings of ACM CCS*, 2018.

[3] Feist, J., Grieco, G., and Groce, A. Slither: A Static Analysis Framework for Smart Contracts. In *IEEE/ACM WETSEB*, 2019.

[4] Durieux, T., Ferreira, J. F., Abreu, R., and Cruz, P. Empirical Review of Automated Analysis Tools on 47,587 Ethereum Smart Contracts. In *Proceedings of ICSE*, 2020.

[5] Liu, Z., Qian, P., Wang, X., Zhu, L., He, Q., and Ji, S. Smart Contract Vulnerability Detection: From Pure Neural Network to Interpretable Graph Feature and Expert Pattern Fusion. In *Proceedings of IJCAI*, 2021.

[6] Nguyen, H. H., et al. MANDO: Multi-Level Heterogeneous Graph Embeddings for Fine-Grained Detection of Smart Contract Vulnerabilities. In *Proceedings of DSAA*, 2022.

[7] Sendner, C., Zhang, R., Hefter, A., Dmitrienko, A., and Koushanfar, F. G-Scan: Graph Neural Networks for Line-Level Vulnerability Identification in Smart Contracts. *arXiv:2307.08549*, 2023.

[8] Jeon, S., Lee, G., Kim, H., and Woo, S. S. Design and Evaluation of Highly Accurate Smart Contract Code Vulnerability Detection Framework. *Data Mining and Knowledge Discovery*, 38, 888-912, 2024.

[9] Salzano, F., et al. An Empirical Analysis of Vulnerability Detection Tools for Solidity Smart Contracts. *Empirical Software Engineering*, 31, 143, 2026.

[10] Feng, Z., Guo, D., Tang, D., et al. CodeBERT: A Pre-Trained Model for Programming and Natural Languages. In *Findings of EMNLP*, 2020.

[11] Guo, D., Ren, S., Lu, S., et al. GraphCodeBERT: Pre-training Code Representations with Data Flow. In *Proceedings of ICLR*, 2021.

[12] Wang, Y., Wang, W., Joty, S., and Hoi, S. C. H. CodeT5: Identifier-aware Unified Pre-trained Encoder-Decoder Models for Code Understanding and Generation. In *Proceedings of EMNLP*, 2021.

[13] Guo, D., Lu, S., Duan, N., Wang, Y., Zhou, M., and Yin, J. UniXcoder: Unified Cross-Modal Pre-training for Code Representation. In *Proceedings of ACL*, 2022.

[14] Roziere, B., Gehring, J., Gloeckle, F., et al. Code Llama: Open Foundation Models for Code. *arXiv:2308.12950*, 2023.

[15] Hui, B., Yang, J., Cui, Z., et al. Qwen2.5-Coder Technical Report. *arXiv preprint*, 2024.

[16] Guo, D., Zhu, Q., Yang, D., et al. DeepSeek-Coder: When the Large Language Model Meets Programming. *arXiv preprint*, 2024.

[17] Lozhkov, A., Li, R., Allal, L. B., et al. StarCoder 2 and The Stack v2: The Next Generation. *arXiv preprint*, 2024.

[18] Team, G. CodeGemma: Open Code Models Based on Gemma. *arXiv preprint*, 2024.

[19] Mishra, M., et al. Granite Code Models: A Family of Open Foundation Models for Code Intelligence. *arXiv preprint*, 2024.

[20] Hu, E. J., Shen, Y., Wallis, P., et al. LoRA: Low-Rank Adaptation of Large Language Models. In *Proceedings of ICLR*, 2022.

[21] Dettmers, T., Pagnoni, A., Holtzman, A., and Zettlemoyer, L. QLoRA: Efficient Finetuning of Quantized LLMs. In *Proceedings of NeurIPS*, 2023.

[22] David, I., et al. Detect Llama: Finding Vulnerabilities in Smart Contracts using Large Language Models. *arXiv:2407.08969*, 2024.

[23] Luu, T. K., Trung, D. M., Tran, T. D., and Duy, P. T. The Fire Tries Gold: Evaluating Pre-trained Language Models for Multi-label Vulnerability Detection in Ethereum Smart Contracts. *Journal of Systems and Software*, 2025.

[24] Bu, et al. SmartBugBERT: BERT-Enhanced Vulnerability Detection for Smart Contract Bytecode. *arXiv:2504.05002*, 2025.
"""

    DRAFT_PATH.write_text(draft, encoding="utf-8")
    HIGHLIGHTS_PATH.write_text(highlights, encoding="utf-8")

    booktabs = "\n".join(
        [
            "% Auto-generated booktabs tables for the Q1 manuscript.",
            r"% Requires \usepackage{booktabs}.",
            "",
            latex_table("三类任务定义与评价指标", "tab:task-definition", task_headers, task_rows, "lllll"),
            latex_table("不同运行模式的平均性能", "tab:mode-means", method_headers, method_table_rows, "lrrrrrr"),
            latex_table("三类任务最佳结果概览", "tab:best-results", best_headers, best_table_rows, "lllrlllll"),
            latex_table("漏洞类型识别 Top 结果", "tab:type-top", type_headers, type_rows, "llrrrrr"),
            latex_table("漏洞行定位 Top 结果", "tab:line-top", line_headers, line_top_rows, "lllrrrr"),
            latex_table("结构化 Prompt 消融结果", "tab:prompt-ablation", prompt_headers, prompt_rows, "llllllll"),
            latex_table("数据补齐前后配对结果", "tab:data-completion", buqi_headers, buqi_rows, "llrrrrr"),
            latex_table("代表性文献数值锚点", "tab:literature-anchors", literature_headers, literature_rows, "llll"),
        ]
    )
    BOOKTABS_PATH.write_text(booktabs, encoding="utf-8")

    summary_md = f"""# Q1 数据表与图表摘要

## 正式矩阵闭合

- 正式实验条目：{formal_total}
- 已有结果：{formal_done}
- 缺口：{formal_missing}
- 模型数：{model_count}
- 模型系列数：{family_count}

## 核心表格

### 任务定义
{task_table}

### 模式均值
{method_table}

### 最佳结果
{best_table}

### 漏洞类型 Top
{type_table}

### 漏洞定位 Top
{line_table}

### Prompt 消融
{prompt_table}

### 补齐前后
{buqi_table}

### 文献数值锚点
{literature_table}

## 图表文件

- {fig('fig1_framework.svg')}
- {fig('fig2_method_performance.svg')}
- {fig('fig3_mode_radar.svg')}
- {fig('fig4_qwen_scaling.svg')}
- {fig('fig5_prompt_ablation.svg')}
- {fig('fig6_data_completion_slope.svg')}
- {fig('fig7_line_boundary_scatter.svg')}
- {fig('fig8_family_mode_heatmap.svg')}
- {fig('fig9_literature_anchor.svg')}
"""
    SUMMARY_PATH.write_text(summary_md, encoding="utf-8")
    print(f"Wrote {DRAFT_PATH}")
    print(f"Wrote {SUMMARY_PATH}")
    print(f"Wrote {HIGHLIGHTS_PATH}")
    print(f"Wrote {BOOKTABS_PATH}")
    print(f"Wrote figures to {FIG_DIR}")


if __name__ == "__main__":
    main()
